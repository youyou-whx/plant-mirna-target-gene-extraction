"""
Auto Blacklist / Whitelist Updater

Reads LLM-reviewed Excel results, collects non-gene and confirmed-gene candidates,
uses LLM batch judgment to decide which entries to add to the blacklist or whitelist,
then writes them back into extract_mirna_genes.py.
"""

import json, os, sys, time, shutil, re
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ============================================================
# Configuration
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LLM_REVIEWED_XLSX = os.path.join(SCRIPT_DIR, "miRNA-Target_Gene_Pairs_Reviewed.xlsx")
FALLBACK_XLSX = os.path.join(SCRIPT_DIR, "miRNA-Target_Gene_Pairs.xlsx")
EXTRACT_PY = os.path.join(SCRIPT_DIR, "extract_mirna_genes.py")

# LLM
LLM_API_URL = "https://api.deepseek.com/v1/chat/completions"
LLM_MODEL = "deepseek-chat"
BATCH_SIZE = 60  # max names per batch for LLM judgment


# ============================================================
# 1. Collect candidates
# ============================================================
def collect_candidates(filepath):
    """Return (non_gene_candidates, whitelist_candidates)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['miRNA-Target Gene Pairs']

    # Detect whether LLM review column exists
    has_llm = 'LLM Review Result' in [str(ws.cell(row=1, column=c).value or '')
                                      for c in range(1, ws.max_column + 1)]

    if not has_llm:
        print("[ERROR] Excel does not have an 'LLM Review Result' column. Run llm_review.py first.")
        wb.close()
        sys.exit(1)

    # A. Non-gene candidates
    non_gene_group = defaultdict(lambda: {"count": 0, "reasons": [], "pmids": set()})
    # B. Whitelist candidates (Confirmed + originally "Unverified" from regex)
    whitelist_group = defaultdict(lambda: {"count": 0, "pmids": set(), "relations": []})

    for row in range(2, ws.max_row + 1):
        gene = str(ws.cell(row=row, column=3).value or '').strip()
        conf = str(ws.cell(row=row, column=4).value or '').strip()
        pmid = str(ws.cell(row=row, column=8).value or '').strip()
        llm_result = str(ws.cell(row=row, column=11).value or '').strip()
        llm_reason = str(ws.cell(row=row, column=12).value or '').strip()

        if not gene:
            continue

        # Non-gene
        if 'Excluded (non-gene)' in llm_result or 'non-gene' in llm_result:
            ng = non_gene_group[gene]
            ng["count"] += 1
            ng["pmids"].add(pmid)
            if llm_reason and llm_reason != 'None':
                # Keep the most concise reason
                reason_short = llm_reason[:150]
                if reason_short not in ng["reasons"]:
                    ng["reasons"].append(reason_short)

        # Whitelist candidate: LLM Confirmed + regex stage confidence was "Unverified"
        if 'Confirmed' in llm_result and 'non-gene' not in llm_result:
            if 'Unverified' in conf:
                wl = whitelist_group[gene]
                wl["count"] += 1
                wl["pmids"].add(pmid)
                # Extract relation type
                rel = llm_result.replace('Confirmed (', '').replace(')', '').strip()
                if rel and rel not in wl["relations"]:
                    wl["relations"].append(rel)

    wb.close()

    # Filter whitelist candidates: must appear at least 2 times
    whitelist_candidates = {k: v for k, v in whitelist_group.items() if v["count"] >= 2}

    # Remove entries already present in existing lists
    existing_blacklist, existing_whitelist = _read_existing_lists()

    non_gene_candidates = {k: v for k, v in non_gene_group.items()
                           if k not in existing_blacklist}
    whitelist_candidates = {k: v for k, v in whitelist_candidates.items()
                            if k not in existing_whitelist}

    print(f"  Non-gene candidates: {len(non_gene_candidates)} unique names (after dedup)")
    print(f"  Whitelist candidates: {len(whitelist_candidates)} unique names (>=2 confirmations, after dedup)")

    return non_gene_candidates, whitelist_candidates


def _read_existing_lists():
    """Read existing BLACKLIST and GENE_WHITELIST from extract_mirna_genes.py"""
    if not os.path.exists(EXTRACT_PY):
        return set(), set()

    with open(EXTRACT_PY, 'r', encoding='utf-8') as f:
        content = f.read()

    blacklist = set()
    whitelist = set()

    # Extract BLACKLIST entries (names in quotes)
    in_blacklist = False
    for line in content.split('\n'):
        stripped = line.strip()
        if 'BLACKLIST' in stripped and '{' in stripped:
            in_blacklist = True
            continue
        if in_blacklist:
            if stripped == '}' or stripped.startswith('}'):
                in_blacklist = False
                continue
            for m in re.finditer(r'"([A-Za-z0-9._+\-]+)"', stripped):
                blacklist.add(m.group(1))

    in_whitelist = False
    for line in content.split('\n'):
        stripped = line.strip()
        if 'GENE_WHITELIST' in stripped and '{' in stripped:
            in_whitelist = True
            continue
        if in_whitelist:
            if stripped == '}' or stripped.startswith('}'):
                in_whitelist = False
                continue
            for m in re.finditer(r'"([A-Za-z0-9._+\-]+)"', stripped):
                whitelist.add(m.group(1))

    return blacklist, whitelist


# ============================================================
# 2. Batch LLM calls
# ============================================================
def call_llm_batch(prompt, api_key):
    """Single LLM call, returns parsed list of dicts"""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 2000,
    }

    for attempt in range(3):
        try:
            import requests as req
            resp = req.post(LLM_API_URL, headers=headers, json=payload, timeout=60)

            if resp.status_code == 200:
                data = resp.json()
                content = data['choices'][0]['message']['content'].strip()
                if '```' in content:
                    content = content.split('```')[1]
                    if content.startswith('json'):
                        content = content[4:]
                    content = content.strip()
                result = json.loads(content)
                return result.get('results', [])
            elif resp.status_code == 429:
                wait = (attempt + 1) * 5
                print(f"   [WARN] Rate limited, retrying in {wait}s...")
                time.sleep(wait)
            else:
                print(f"   [WARN] HTTP {resp.status_code}: {resp.text[:100]}")
                time.sleep(2)
        except Exception as e:
            print(f"   [WARN] Error: {e}")
            time.sleep(2)

    return []


def judge_blacklist(candidates, api_key):
    """Batch judge which names should be added to the blacklist"""
    if not candidates:
        return []

    # Build data list
    items = []
    for name, info in sorted(candidates.items(), key=lambda x: -x[1]["count"]):
        reasons = '; '.join(info['reasons'][:2])
        items.append({
            "name": name,
            "count": info["count"],
            "reasons": reasons,
        })

    # Process in batches
    all_results = []
    for batch_start in range(0, len(items), BATCH_SIZE):
        batch = items[batch_start:batch_start + BATCH_SIZE]
        batch_json = json.dumps(batch, ensure_ascii=False, indent=2)

        prompt = f"""你是植物分子生物学专家。以下名称被 LLM 审查标记为"非基因"。
请判断每个名称是否应该加入黑名单（在以后的提取中直接排除）。

⚠️ 铁律：已知的植物基因家族名绝不加入黑名单！
  即使某篇论文中它恰好不是基因（如作为蛋白域名出现），也不能黑名单。
  已知植物基因家族包括但不限于：NAC, WRKY, MYB, TCP, SPL, AGO, AP2/ERF, ARF,
  GRF, NAC, DCL, RDR, bZIP, bHLH, MADS, HD-ZIP, NLR, PPR, PCF, LRR, RLK 等。
  这些名称在植物学文献中极为常见，即使偶尔作为非基因出现也是语境依赖，不应黑名单。

黑名单标准：这个名称在任何水稻/植物论文中都不太可能是基因名。
  应加入（always_non_gene）：
    - 品种名/栽培种（如 IR56, ZH11, N22）
    - 菌株名/分离株（如 PXO86, Guy11）
    - 纯生物学过程缩写（如 PTI=PAMP-triggered immunity, ETI, SAR）
    - 植物激素/化合物（如 GA, ABA, JA, SA）
    - 实验技术/工具（如 PCR, GFP, ChIP）
    - 人类/动物基因名
    - 截断词（来自一个更长的词但不是独立基因，如 LIKE 来自 DICER-LIKE1）
  不应加入（context_dependent）：
    - 名称可能是基因/基因家族名，只是在某篇论文中恰好不是
    - 判定理由明确说了"本句中/该文中"等语境依赖表述
    - 名称是已知的植物基因家族名或转录因子家族名

数据（共 {len(batch)} 个）：
{batch_json}

输出纯JSON（不要markdown包裹）：
{{"results": [{{"name": "PTI", "decision": "always_non_gene", "brief": "免疫过程缩写"}}, ...]}}"""

        print(f"   [LLM] Batch judging blacklist {batch_start+1}-{min(batch_start+BATCH_SIZE, len(items))}/{len(items)} ...")
        results = call_llm_batch(prompt, api_key)
        all_results.extend(results)
        if batch_start + BATCH_SIZE < len(items):
            time.sleep(0.5)

    return all_results


def judge_whitelist(candidates, api_key):
    """Batch judge which names should be added to the whitelist"""
    if not candidates:
        return []

    items = []
    for name, info in sorted(candidates.items(), key=lambda x: -x[1]["count"]):
        items.append({
            "name": name,
            "count": info["count"],
            "pmid_count": len(info["pmids"]),
            "relations": ', '.join(info["relations"][:3]),
        })

    all_results = []
    for batch_start in range(0, len(items), BATCH_SIZE):
        batch = items[batch_start:batch_start + BATCH_SIZE]
        batch_json = json.dumps(batch, ensure_ascii=False, indent=2)

        prompt = f"""你是植物分子生物学专家。以下名称在正则初筛时被标为"待验证"，但经 LLM 审查确认为真实植物基因。
请判断哪些应该加入基因白名单（以后直接标为高置信）。

白名单标准：这是一个规范的植物/水稻基因名。
  应加入（add）：
    - 符合植物基因命名规范（大写字母+数字，如 NRAMP3, SIP19）
    - 已知的植物基因家族成员
    - 有明确靶向关系支持
  不应加入（skip）：
    - 名称太泛，可能在不同语境下指代不同东西
    - 纯数字编号或无意义缩写
    - 只在单篇文献中出现（证据不够强）

数据（共 {len(batch)} 个）：
{batch_json}

输出纯JSON（不要markdown包裹）：
{{"results": [{{"name": "NRAMP3", "decision": "add", "brief": "水稻重金属转运蛋白基因"}}, ...]}}"""

        print(f"   [LLM] Batch judging whitelist {batch_start+1}-{min(batch_start+BATCH_SIZE, len(items))}/{len(items)} ...")
        results = call_llm_batch(prompt, api_key)
        all_results.extend(results)
        if batch_start + BATCH_SIZE < len(items):
            time.sleep(0.5)

    return all_results


# ============================================================
# 3. Write back to extract_mirna_genes.py
# ============================================================
def apply_updates(blacklist_add, whitelist_add):
    """Insert new entries into extract_mirna_genes.py"""
    if not blacklist_add and not whitelist_add:
        print("\n  No new entries to add.")
        return

    # Backup
    backup_path = EXTRACT_PY + ".bak"
    shutil.copy2(EXTRACT_PY, backup_path)
    print(f"\n  Backup created: {backup_path}")

    with open(EXTRACT_PY, 'r', encoding='utf-8') as f:
        content = f.read()

    existing_blacklist, existing_whitelist = _read_existing_lists()

    # Dedup + safety valve: blacklist must never overwrite whitelist
    blacklist_new = sorted(set(blacklist_add) - existing_blacklist - existing_whitelist)
    whitelist_new = sorted(set(whitelist_add) - existing_whitelist - existing_blacklist)

    print(f"  Blacklist additions: {len(blacklist_new)} entries → {blacklist_new}")
    print(f"  Whitelist additions: {len(whitelist_new)} entries → {whitelist_new}")

    if blacklist_new:
        insert_str = '\n'.join(f'    "{name}",' for name in blacklist_new)
        # Find the insertion marker in the blacklist section
        marker = '# ====== Added after first round ======'
        if marker in content:
            # Insert before the last "}" in the blacklist, after the last existing entry
            pattern = r'(# ----- Other false positives.*?\n\s*"[^"]*".*?\n)'
            match = re.search(pattern, content, re.DOTALL)
            if match:
                insert_point = match.end()
                content = (content[:insert_point]
                          + f'\n    # ----- LLM auto-added ({len(blacklist_new)} entries) -----\n'
                          + insert_str + '\n'
                          + content[insert_point:])

    if whitelist_new:
        insert_str = '\n'.join(f'    "{name}",' for name in whitelist_new)
        # Find the closing "}" of GENE_WHITELIST
        whitelist_start = content.find('GENE_WHITELIST')
        if whitelist_start >= 0:
            brace_count = 0
            in_set = False
            insert_point = whitelist_start
            for i in range(whitelist_start, len(content)):
                if content[i] == '{':
                    brace_count += 1
                    in_set = True
                elif content[i] == '}':
                    brace_count -= 1
                    if in_set and brace_count == 0:
                        insert_point = i
                        break
            if insert_point > whitelist_start:
                content = (content[:insert_point]
                          + f'\n    # ----- LLM auto-added ({len(whitelist_new)} entries) -----\n'
                          + insert_str + '\n'
                          + content[insert_point:])

    with open(EXTRACT_PY, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f"   [OK] Written to {EXTRACT_PY}")


# ============================================================
# 4. Main
# ============================================================
def main():
    apply_flag = '--apply' in sys.argv

    # Check API Key
    api_key = os.environ.get('DEEPSEEK_API_KEY', '')
    if not api_key:
        print("=" * 60)
        print("  ERROR: DEEPSEEK_API_KEY not set")
        print('  Run: $env:DEEPSEEK_API_KEY = "sk-xxx"')
        print("=" * 60)
        sys.exit(1)

    # Determine input file
    if os.path.exists(LLM_REVIEWED_XLSX):
        fpath = LLM_REVIEWED_XLSX
        print(f"[INFO] Reading LLM review results: {os.path.basename(fpath)}")
    elif os.path.exists(FALLBACK_XLSX):
        fpath = FALLBACK_XLSX
        print(f"[WARN] LLM review results not found, using regex extraction results: {os.path.basename(fpath)}")
    else:
        print("[ERROR] No input file found!")
        sys.exit(1)

    # 1. Collect candidates
    print("\n[INFO] Collecting candidates...")
    non_gene_candidates, whitelist_candidates = collect_candidates(fpath)

    if not non_gene_candidates and not whitelist_candidates:
        print("\n   No new candidates — exiting.")
        return

    # 2. LLM batch judgment
    blacklist_add = []
    whitelist_add = []

    if non_gene_candidates:
        print(f"\n[INFO] LLM judging {len(non_gene_candidates)} non-gene candidates...")
        results = judge_blacklist(non_gene_candidates, api_key)
        for r in results:
            if r.get('decision') == 'always_non_gene':
                blacklist_add.append(r['name'])
                print(f"   [BL] {r['name']:20s} → blacklist  ({r.get('brief', '')})")
            else:
                print(f"   [SKIP] {r['name']:20s} → skip      ({r.get('brief', '')})")

    if whitelist_candidates:
        print(f"\n[INFO] LLM judging {len(whitelist_candidates)} whitelist candidates...")
        results = judge_whitelist(whitelist_candidates, api_key)
        for r in results:
            if r.get('decision') == 'add':
                whitelist_add.append(r['name'])
                print(f"   [WL] {r['name']:20s} → whitelist  ({r.get('brief', '')})")
            else:
                print(f"   [SKIP] {r['name']:20s} → skip      ({r.get('brief', '')})")

    # 3. Write
    if apply_flag:
        print("\n[INFO] Writing to extract_mirna_genes.py ...")
        apply_updates(blacklist_add, whitelist_add)
    else:
        print(f"\n{'='*55}")
        print(f"  Preview complete!")
        print(f"  Blacklist additions: {len(blacklist_add)}")
        print(f"  Whitelist additions: {len(whitelist_add)}")
        print(f"")
        print(f"  To apply changes, run:")
        print(f"    python auto_update_lists.py --apply")
        print(f"{'='*55}")


if __name__ == "__main__":
    main()
