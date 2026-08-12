"""
miRNA Normalization, Deduplication & Gene ID Resolution

Input:  LLM-reviewed Excel (or regex extraction Excel)
Output: Normalized, deduplicated Excel with gene database IDs (RAP, MSU, NCBI)

Usage:
    python normalize_names.py [input_file.xlsx]
"""

import re
import sys
import os
import time
import urllib.request
import urllib.parse
import urllib.error
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ============================================================
# Configuration
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT = os.path.join(SCRIPT_DIR, "miRNA-Target_Gene_Pairs_Reviewed.xlsx")
FALLBACK_INPUT = os.path.join(SCRIPT_DIR, "miRNA-Target_Gene_Pairs.xlsx")

# ============================================================
# Species prefix mapping (2-letter uppercase → 3-letter lowercase standard)
# ============================================================
SPECIES_PREFIX_MAP = {
    # Major cereals
    "Os": "osa",    # Oryza sativa (rice)
    "Ta": "tae",    # Triticum aestivum (wheat)
    "Zm": "zma",    # Zea mays (maize)
    "Hvu": "hvu",   # Hordeum vulgare (barley)
    "Sbi": "sbi",   # Sorghum bicolor
    # Model plants
    "At": "ath",    # Arabidopsis thaliana
    "Ath": "ath",   # Arabidopsis thaliana (variant)
    "Al": "aly",    # Arabidopsis lyrata
    # Legumes
    "Gm": "gma",    # Glycine max (soybean)
    "Pv": "pvu",    # Phaseolus vulgaris (common bean)
    "Mt": "mtr",    # Medicago truncatula (alfalfa)
    "Cas": "cas",   # Cicer arietinum (chickpea)
    "Lj": "lja",    # Lotus japonicus
    # Solanaceae
    "Sly": "sly",   # Solanum lycopersicum (tomato)
    "Stu": "stu",   # Solanum tuberosum (potato)
    "Nta": "nta",   # Nicotiana tabacum (tobacco)
    "Ca": "can",    # Capsicum annuum (pepper)
    # Fruits
    "Ppe": "ppe",   # Prunus persica (peach)
    "Mdm": "mdm",   # Malus domestica (apple)
    "Vvi": "vvi",   # Vitis vinifera (grape)
    "Csi": "csi",   # Citrus sinensis (orange)
    "Cm": "cme",    # Cucumis melo (melon/cucumber)
    "Fv": "far",    # Fragaria vesca (strawberry)
    "Ma": "mac",    # Musa acuminata (banana)
    "Cp": "cpa",    # Carica papaya (papaya)
    # Fiber / oil crops
    "Gh": "ghr",    # Gossypium hirsutum (cotton)
    "Bna": "bna",   # Brassica napus (rapeseed)
    "Bra": "bra",   # Brassica rapa (turnip)
    # Trees
    "Ptc": "ptc",   # Populus trichocarpa (poplar)
    "Pt": "pta",    # Pinus taeda (pine)
    "Mes": "mes",   # Manihot esculenta (cassava)
}

# Known standard 3-letter prefixes (no conversion needed)
STD_PREFIXES = set(SPECIES_PREFIX_MAP.values())

# Species table for resolving defaults (mirrored from extract_mirna_genes.py)
_SPECIES_TABLE = {
    # Major cereals
    "osa": ("Os", "rice"),
    "tae": ("Ta", "wheat"),
    "zma": ("Zm", "maize"),
    "hvu": ("Hvu", "barley"),
    "sbi": ("Sbi", "sorghum"),
    # Model plants
    "ath": ("At", "Arabidopsis"),
    "aly": ("Al", "Arabidopsis lyrata"),
    # Legumes
    "gma": ("Gm", "soybean"),
    "pvu": ("Pv", "common bean"),
    "mtr": ("Mt", "alfalfa"),
    "cas": ("Cas", "chickpea"),
    "lja": ("Lj", "lotus"),
    # Solanaceae
    "sly": ("Sly", "tomato"),
    "stu": ("Stu", "potato"),
    "nta": ("Nta", "tobacco"),
    "can": ("Ca", "pepper"),
    # Fruits
    "ppe": ("Ppe", "peach"),
    "mdm": ("Mdm", "apple"),
    "vvi": ("Vvi", "grape"),
    "csi": ("Csi", "orange"),
    "cme": ("Cm", "melon"),
    "far": ("Fv", "strawberry"),
    "mac": ("Ma", "banana"),
    "cpa": ("Cp", "papaya"),
    # Fiber / oil crops
    "ghr": ("Gh", "cotton"),
    "bna": ("Bna", "rapeseed"),
    "bra": ("Bra", "turnip"),
    # Trees
    "ptc": ("Ptc", "poplar"),
    "pta": ("Pt", "pine"),
    "mes": ("Mes", "cassava"),
}


def _species_table():
    return _SPECIES_TABLE


def normalize_mirna(name, species=None):
    """
    Normalize miRNA names:
      OsmiR156     → osa-miR156
      OsmiR156a    → osa-miR156a
      OsmiR156a-5p → osa-miR156a-5p
      miR156       → osa-miR156       (no species prefix, uses `species` param or defaults to rice)
      miR156a-5p   → osa-miR156a-5p
      osa-miR156a  → osa-miR156a      (already standard, keep)
      pre-miR156   → osa-miR156       (precursor → mature)
      pri-miR408   → osa-miR408
      ath-miR843   → ath-miR843       (non-rice, keep as-is)

    `species` is the common name (e.g. "rice", "wheat") used to resolve bare names.
    """
    name = name.strip()

    # Resolve default miRNA prefix from species
    default_mir_prefix = None  # if unknown, bare names stay bare
    if species and species != "unknown":
        for mir_key, (_gene_pfx, common) in _species_table().items():
            if common == species:
                default_mir_prefix = mir_key
                break

    # ── Remove precursor tags pre-/pri- ──
    is_precursor = False
    for tag in ['pre-', 'pri-']:
        if name.lower().startswith(tag):
            name = name[len(tag):]
            is_precursor = True
            break

    # ── Case 1: already standard format xxx-miR... ──
    m_std = re.match(r'^([a-z]{3,5})-[Mm][Ii][Rr](\d+)([a-z]*)([.-]\d[ap]?)?$', name)
    if m_std:
        return name

    # ── Case 2: legacy species prefix OsmiR156 / TamiR156 ──
    m_old = re.match(r'^([A-Z][a-z]?)([Mm][Ii][Rr]\d+[a-z]*(?:[-.]\d[ap]?)?)$', name)
    if m_old:
        prefix = m_old.group(1)
        rest = m_old.group(2)
        if prefix in SPECIES_PREFIX_MAP:
            return f"{SPECIES_PREFIX_MAP[prefix]}-{rest}"
        return name

    # ── Case 3: no species prefix miR156 / miR156a-5p ──
    m_bare = re.match(r'^[Mm][Ii][Rr](\d+)([a-z]*)([.-]\d[ap]?)?$', name)
    if m_bare:
        if default_mir_prefix is None:
            return name  # species unknown — keep bare, don't guess
        num = m_bare.group(1)
        letter = m_bare.group(2) or ''
        arm = m_bare.group(3) or ''
        return f"{default_mir_prefix}-miR{num}{letter}{arm}"

    # ── Case 4: mirtron or other special forms ──
    m_multi = re.match(r'^[Mm][Ii][Rr](\d+)([a-z]+)([.-]\d[ap]?)?$', name)
    if m_multi:
        if default_mir_prefix is None:
            return name  # species unknown — keep as-is
        num = m_multi.group(1)
        letters = m_multi.group(2)
        arm = m_multi.group(3) or ''
        return f"{default_mir_prefix}-miR{num}{letters}{arm}"

    # ── Case 5: unrecognizable → keep as-is ──
    return name


# ============================================================
# Gene names — kept as-is (normalization deferred)
# ============================================================

def normalize_gene(name, species=None):
    """Gene names are preserved in their original extracted form."""
    return name.strip()


# ============================================================
# Read Excel
# ============================================================
def load_pairs(filepath):
    """
    Load pair data from Excel.
    Auto-detects: regex extraction output (9 columns) vs. LLM review output (11 columns).
    Only retains "Confirmed" entries; excludes non-gene, non-target, and weak-association.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['miRNA-Target Gene Pairs']

    # Detect column structure
    headers = [str(ws.cell(row=1, column=c).value or '') for c in range(1, ws.max_column + 1)]

    has_llm_review = 'LLM Review Result' in headers

    pairs = []
    skipped = {"non-gene": 0, "non-target": 0, "weak": 0, "not_reviewed": 0}

    for row in range(2, ws.max_row + 1):
        mirna = str(ws.cell(row=row, column=2).value or '').strip()
        gene = str(ws.cell(row=row, column=3).value or '').strip()
        conf = str(ws.cell(row=row, column=4).value or '').strip()
        rel = str(ws.cell(row=row, column=5).value or '').strip()
        sent = str(ws.cell(row=row, column=6).value or '').strip()
        title = str(ws.cell(row=row, column=7).value or '').strip()
        pmid = str(ws.cell(row=row, column=8).value or '').strip()
        doi = str(ws.cell(row=row, column=9).value or '').strip()
        species = str(ws.cell(row=row, column=10).value or '').strip()

        if not mirna or not gene:
            continue

        # If LLM review column exists, filter (shifted to col 11 due to Species column)
        if has_llm_review:
            llm_result = str(ws.cell(row=row, column=11).value or '').strip()
            keep_mark = str(ws.cell(row=row, column=13).value or '').strip()
            # Confirmed + pre-filled "Yes" → always keep
            # Weak/other + user wrote "Yes" → keep
            manual_keep = keep_mark.lower() in ("yes", "y", "keep", "1", "true")
            if llm_result.startswith("Confirmed") or manual_keep:
                pass  # keep this row
            elif 'Weak association' in llm_result:
                skipped["weak"] += 1
                continue
            elif 'Excluded' in llm_result or 'non-gene' in llm_result:
                if 'non-target' in llm_result:
                    skipped["non-target"] += 1
                else:
                    skipped["non-gene"] += 1
                continue
            elif 'Not reviewed' in llm_result:
                skipped["not_reviewed"] += 1
                continue
            else:
                skipped["not_reviewed"] += 1
                continue
        else:
            # No LLM review → retain all (regex extraction only)
            pass

        pairs.append({
            "mirna_raw": mirna,
            "gene_raw": gene,
            "confidence": conf,
            "relation": rel,
            "sentence": sent,
            "title": title,
            "pmid": pmid,
            "doi": doi,
            "species": species,
        })

    wb.close()

    if has_llm_review:
        print(f"   Loaded {len(pairs)} entries "
              f"(skipped {skipped['non-gene']} non-gene, "
              f"{skipped['non-target']} non-target, "
              f"{skipped['weak']} weak, "
              f"{skipped['not_reviewed']} not-reviewed)")
    else:
        print(f"   Loaded {len(pairs)} entries (no LLM review column, all retained)")

    return pairs


# ============================================================
# Deduplication & merge
# ============================================================
def deduplicate(pairs):
    """
    Same (normalized miRNA, normalized gene) → keep one row,
    merge all source sentences, PMIDs, DOIs, titles.
    """
    groups = defaultdict(lambda: {
        "mirna_norm": "",
        "gene_norm": "",
        "mirna_variants": set(),
        "gene_variants": set(),
        "sentences": [],
        "titles": [],
        "pmids": [],
        "dois": [],
        "confidences": [],
        "relations": [],
    })

    for p in pairs:
        m_norm = p["mirna_norm"]
        g_norm = p["gene_norm"]
        species = p.get("species", "unknown")
        key = (species, m_norm, g_norm)

        g = groups[key]
        g["mirna_norm"] = m_norm
        g["gene_norm"] = g_norm
        g["mirna_variants"].add(p["mirna_raw"])
        g["gene_variants"].add(p["gene_raw"])

        if p["sentence"] and p["sentence"] not in g["sentences"]:
            g["sentences"].append(p["sentence"])
        if p["title"] and p["title"] not in g["titles"]:
            g["titles"].append(p["title"])
        if p["pmid"] and p["pmid"] not in g["pmids"]:
            g["pmids"].append(p["pmid"])
        if p["doi"] and p["doi"] not in g["dois"]:
            g["dois"].append(p["doi"])
        if p["confidence"]:
            g["confidences"].append(p["confidence"])
        if p["relation"]:
            g["relations"].append(p["relation"])

    # Build output
    result = []
    for key, g in sorted(groups.items()):
        # Show variants
        m_display = g["mirna_norm"]
        if g["mirna_variants"] != {g["mirna_norm"]}:
            variants = g["mirna_variants"] - {g["mirna_norm"]}
            if variants:
                m_display += " (" + ", ".join(sorted(variants)) + ")"

        g_display = g["gene_norm"]
        if g["gene_variants"] != {g["gene_norm"]}:
            variants = g["gene_variants"] - {g["gene_norm"]}
            if variants:
                g_display += " (" + ", ".join(sorted(variants)) + ")"

        # Merge sentences (deduped, grouped by PMID)
        sentences_combined = " | ".join(g["sentences"])

        # Merge PMID / DOI
        pmids_combined = ", ".join(g["pmids"])
        dois_combined = ", ".join(g["dois"])
        titles_combined = " | ".join(g["titles"])

        # Most common relation type
        rel = max(set(g["relations"]), key=g["relations"].count) if g["relations"] else ""

        result.append({
            "mirna": m_display,
            "gene": g_display,
            "mirna_norm": g["mirna_norm"],
            "gene_norm": g["gene_norm"],
            "species": key[0],
            "pmid_count": len(g["pmids"]),
            "relation": rel,
            "sentences": sentences_combined,
            "titles": titles_combined,
            "pmids": pmids_combined,
            "dois": dois_combined,
        })

    return result


# ============================================================
# Gene ID Resolution (ricedata.cn for rice, NCBI E-utils for others)
# ============================================================

RICEDATA_URL = "https://www.ricedata.cn/gene/accessions_switch.aspx"
NCBI_ESEARCH = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"

# Common name → NCBI organism name
_NCBI_ORGANISM = {
    "rice": "Oryza sativa", "wheat": "Triticum aestivum",
    "maize": "Zea mays", "Arabidopsis": "Arabidopsis thaliana",
    "soybean": "Glycine max", "barley": "Hordeum vulgare",
    "sorghum": "Sorghum bicolor", "tomato": "Solanum lycopersicum",
    "potato": "Solanum tuberosum", "tobacco": "Nicotiana tabacum",
    "poplar": "Populus trichocarpa", "grape": "Vitis vinifera",
    "peach": "Prunus persica", "apple": "Malus domestica",
    "rapeseed": "Brassica napus", "turnip": "Brassica rapa",
    "cotton": "Gossypium hirsutum", "orange": "Citrus sinensis",
    "cassava": "Manihot esculenta", "common bean": "Phaseolus vulgaris",
    "alfalfa": "Medicago truncatula", "chickpea": "Cicer arietinum",
    "lotus": "Lotus japonicus", "pepper": "Capsicum annuum",
    "melon": "Cucumis melo", "strawberry": "Fragaria vesca",
    "banana": "Musa acuminata", "papaya": "Carica papaya",
    "pine": "Pinus taeda", "Arabidopsis lyrata": "Arabidopsis lyrata",
}


def _lookup_ricedata(gene):
    """Query ricedata.cn for a rice gene. Returns dict or None."""
    url = f"{RICEDATA_URL}?para={urllib.parse.quote(gene)}"
    try:
        req = urllib.request.Request(url)
        req.add_header("User-Agent", "Mozilla/5.0")
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode("gb2312", errors="replace")
    except Exception:
        return None

    # Parse the TBResult table
    table_m = re.search(r"<TABLE[^>]*id='TBResult'[^>]*>(.*?)</TABLE>", html, re.DOTALL)
    if not table_m:
        return None

    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", table_m.group(1), re.DOTALL | re.IGNORECASE)
    all_cells = []
    for row_html in rows:
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.DOTALL | re.IGNORECASE)
        cells = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
        cells = [re.sub(r"&nbsp;", " ", c).strip() for c in cells]
        all_cells.append(cells)

    # Filter header rows, find data row
    skip = ("GeneID", "基因名称", "基因符号", "RAP", "共")
    for cells in all_cells:
        if len(cells) >= 4 and cells[0] and not any(cells[0].startswith(w) for w in skip):
            # Extract first word from RAP/MSU cells (strip ad text)
            def _first(cell):
                m = re.match(r"^(\S+)", cell)
                return m.group(1) if m else cell
            return {
                "RAP": _first(cells[3]) if len(cells) > 3 else "",
                "MSU": _first(cells[4]) if len(cells) > 4 else "",
                "NCBI": _first(cells[5]) if len(cells) > 5 else "",
            }
    return None


def _lookup_ncbi(species, gene):
    """Query NCBI E-utils. Returns Entrez ID or None."""
    org = _NCBI_ORGANISM.get(species, species)

    # Try full name first, then without species prefix (still species-filtered)
    queries = [gene]
    prefix = _NCBI_GENE_PREFIX.get(species, "")
    if prefix and gene.startswith(prefix) and len(gene) > len(prefix) + 1:
        queries.append(gene[len(prefix):])

    for q in queries:
        query = f"{q}[Gene] AND {org}[Organism]"
        params = {"db": "gene", "term": query, "retmode": "json", "retmax": 3}
        url = NCBI_ESEARCH + "?" + urllib.parse.urlencode(params)
        try:
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = __import__("json").loads(resp.read().decode("utf-8"))
            ids = data.get("esearchresult", {}).get("idlist", [])
            if ids:
                return ids[0]
        except Exception:
            pass
    return None


# Gene prefix to strip for NCBI fallback queries
_NCBI_GENE_PREFIX = {
    "rice": "Os", "wheat": "Ta", "maize": "Zm", "Arabidopsis": "At",
    "soybean": "Gm", "barley": "Hvu", "sorghum": "Sbi", "tomato": "Sly",
    "potato": "Stu", "tobacco": "Nta", "poplar": "Ptc", "peach": "Ppe",
    "apple": "Mdm", "grape": "Vvi", "rapeseed": "Bna", "turnip": "Bra",
    "cotton": "Gh", "orange": "Csi", "cassava": "Mes",
    "common bean": "Pv", "alfalfa": "Mt", "chickpea": "Cas",
    "lotus": "Lj", "pepper": "Ca", "melon": "Cm", "strawberry": "Fv",
    "banana": "Ma", "papaya": "Cp", "pine": "Pt",
    "Arabidopsis lyrata": "Al",
}


def resolve_gene_ids(pairs_after, progress_cb=None):
    """
    Resolve gene IDs for all unique genes in deduped pairs.
    Rice → ricedata.cn, others → NCBI E-utils.
    Returns dict {(species, gene_norm): {RAP, MSU, NCBI}}.
    If progress_cb(ratio, text) is provided, it's called after each gene.
    """
    # Collect unique (species, gene) from deduped results
    unique = set()
    for p in pairs_after:
        sp = p.get("species", "unknown")
        gn = p["gene_norm"]
        if sp != "unknown":
            unique.add((sp, gn))

    genes_list = sorted(unique)
    total = len(genes_list)
    mapping = {}
    resolved = 0

    print(f"\n[INFO] Resolving gene IDs for {total} unique genes...")
    rice_count = sum(1 for sp, _ in genes_list if sp == "rice")
    print(f"   Rice (ricedata.cn): {rice_count}  |  Others (NCBI E-utils): {total - rice_count}\n")

    for i, (sp, gn) in enumerate(genes_list):
        if progress_cb:
            progress_cb((i + 1) / total, f"Resolving gene IDs ({i + 1}/{total})")
        if sp == "rice":
            info = _lookup_ricedata(gn)
            if info:
                mapping[(sp, gn)] = info
                resolved += 1
                print(f"  [{i+1:4d}/{total}] + {gn:30s} RAP:{info.get('RAP',''):15s} NCBI:{info.get('NCBI','')}")
            else:
                mapping[(sp, gn)] = {"RAP": "", "MSU": "", "NCBI": ""}
                print(f"  [{i+1:4d}/{total}] - {gn:30s} not in ricedata.cn")
            time.sleep(0.3)
        else:
            eid = _lookup_ncbi(sp, gn)
            if eid:
                mapping[(sp, gn)] = {"RAP": "", "MSU": "", "NCBI": eid}
                resolved += 1
                print(f"  [{i+1:4d}/{total}] + {gn:30s} NCBI:{eid}")
            else:
                mapping[(sp, gn)] = {"RAP": "", "MSU": "", "NCBI": ""}
                print(f"  [{i+1:4d}/{total}] - {gn:30s} not found in NCBI Gene")
            time.sleep(0.4)

    print(f"\n[INFO] Gene ID resolution: {resolved}/{total} ({resolved/total*100:.1f}%)")
    return mapping


# ============================================================
# Excel output
# ============================================================
HEADER_FONT = Font(name='Consolas', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

NORMAL_FONT = Font(name='Consolas', size=10)
BOLD_FONT = Font(name='Consolas', size=10, bold=True)
WRAP_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def export_excel(pairs_before, pairs_after, gene_map, gene_resolved, gene_total, input_file, output_path):
    """
    Output three sheets:
      Sheet 1: Normalized & deduplicated pairs (with gene IDs)
      Sheet 2: Normalized detail (per-sentence granularity)
      Sheet 3: Statistics (extraction → review → normalization → gene IDs)
    """
    wb = openpyxl.Workbook()

    # ========================
    # Sheet 1: Deduplicated & merged
    # ========================
    ws1 = wb.active
    ws1.title = "Normalized-Deduped"

    headers1 = ["#", "Normalized miRNA", "Gene", "Species", "RAP_ID", "MSU_ID",
                "NCBI_GeneID", "PMID Count", "Relation Type",
                "Source Sentences (merged)", "Article Titles", "PMIDs", "DOIs"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for i, p in enumerate(pairs_after):
        row = i + 2
        sp = p.get("species", "unknown")
        gn = p.get("gene_norm", p["gene"])
        gmap = gene_map.get((sp, gn), {})
        vals = [i + 1, p["mirna"], p["gene"], p.get("species", "unknown"),
                gmap.get("RAP", ""), gmap.get("MSU", ""), gmap.get("NCBI", ""),
                p["pmid_count"], p["relation"],
                p["sentences"], p["titles"], p["pmids"], p["dois"]]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT if col in (10, 11) else CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            if p["pmid_count"] >= 2:
                cell.fill = LIGHT_BLUE_FILL

    col_widths1 = [6, 22, 22, 12, 16, 16, 14, 8, 12, 70, 50, 20, 30]
    for col, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(pairs_after) + 1}"

    # ========================
    # Sheet 2: Normalized detail (per-sentence granularity)
    # ========================
    ws2 = wb.create_sheet("Normalized-Detail")
    headers2 = ["#", "Original miRNA", "Normalized miRNA", "Gene",
                "Species", "Gene Confidence", "Relation Type", "Source Sentence", "Article Title", "PMID", "DOI"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for i, p in enumerate(pairs_before):
        row = i + 2
        m_changed = p["mirna_raw"] != p["mirna_norm"]

        vals = [i + 1, p["mirna_raw"], p["mirna_norm"],
                p["gene_raw"],
                p.get("species", "unknown"),
                p["confidence"], p["relation"],
                p["sentence"], p["title"], p["pmid"], p["doi"]]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT if col in (8, 9) else CENTER_ALIGNMENT
            cell.border = THIN_BORDER

        # Highlight miRNA name changes
        if m_changed:
            ws2.cell(row=row, column=3).fill = GREEN_FILL

    col_widths2 = [6, 18, 18, 16, 12, 14, 12, 70, 50, 12, 28]
    for col, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(pairs_before) + 1}"

    # ========================
    # Sheet 3: Statistics (all stages)
    # ========================
    ws3 = wb.create_sheet("Statistics")

    bold12 = Font(name='Consolas', bold=True, size=12, color='2F5496')
    bold11 = Font(name='Consolas', bold=True, size=11)

    title_cell = ws3.cell(row=1, column=1, value="Pipeline Statistics — All Stages")
    title_cell.font = Font(name='Consolas', bold=True, size=14, color='2F5496')
    ws3.merge_cells('A1:C1')

    # Copy extraction + review stats from the input file's Statistics sheet
    row = 3
    try:
        src_wb = openpyxl.load_workbook(input_file)
        if 'Statistics' in src_wb.sheetnames:
            src_ws = src_wb['Statistics']
            for src_row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, values_only=True):
                vals = [str(c) if c is not None else "" for c in src_row]
                if any(v for v in vals):
                    for col, v in enumerate(vals[:3], 1):
                        cell = ws3.cell(row=row, column=col, value=v)
                        if row <= 3 or str(v).startswith("---") or str(v).startswith("["):
                            cell.font = bold12 if str(v).startswith("---") or str(v).startswith("[") else bold11
                        else:
                            cell.font = NORMAL_FONT
                    row += 1
        src_wb.close()
    except Exception:
        pass

    row += 1

    # Normalization + Dedup stats
    m_changed = sum(1 for p in pairs_before if p["mirna_raw"] != p["mirna_norm"])
    unique_m_before = len(set(p["mirna_raw"] for p in pairs_before))
    unique_m_after = len(set(p["mirna_norm"] for p in pairs_before))

    now_stats = [
        ("--- Normalization & Dedup ---", "", ""),
        ("Pairs (before normalization)", len(pairs_before), ""),
        ("  miRNA names changed", m_changed,
         f"{m_changed/len(pairs_before)*100:.1f}%" if pairs_before else ""),
        ("  Gene names — preserved as-is", "", ""),
        ("Unique miRNAs (before)", unique_m_before, f"→ {unique_m_after} (after)"),
        ("", "", ""),
        ("Pair groups (after dedup, before ID resolution)", len(pairs_after),
         f"Merged {len(pairs_before)-len(pairs_after)} entries"),
        ("Multi-PMID support (>=2)", sum(1 for p in pairs_after if p["pmid_count"] >= 2), ""),
        ("", "", ""),
        ("--- Gene ID Resolution ---", "", ""),
        ("Unique genes queried", gene_total, ""),
        ("Resolved (RAP/MSU/NCBI)", gene_resolved, f"{gene_resolved/gene_total*100:.1f}%" if gene_total else ""),
        ("Unresolved", gene_total - gene_resolved, ""),
    ]

    for label, val1, val2 in now_stats:
        if label.startswith("---"):
            ws3.cell(row=row, column=1, value=label).font = bold12
        elif label == "":
            pass
        else:
            ws3.cell(row=row, column=1, value=label).font = NORMAL_FONT
        if val1 != "":
            ws3.cell(row=row, column=2, value=val1).font = NORMAL_FONT
        if val2 != "":
            ws3.cell(row=row, column=3, value=val2).font = NORMAL_FONT
        row += 1

    ws3.column_dimensions['A'].width = 42
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 28

    if os.path.exists(output_path):
        os.remove(output_path)
    wb.save(output_path)
    wb.close()
    print(f"   Excel saved to: {output_path}")


# ============================================================
# Main
# ============================================================
def main(input_file=None):
    # Determine input file
    if input_file:
        fpath = input_file
    elif os.path.exists(DEFAULT_INPUT):
        fpath = DEFAULT_INPUT
        print(f"[INFO] Using LLM review results: {os.path.basename(fpath)}")
    elif os.path.exists(FALLBACK_INPUT):
        fpath = FALLBACK_INPUT
        print(f"[WARN] LLM review results not found, using regex extraction results: {os.path.basename(fpath)}")
    else:
        print("[ERROR] No input file found!")
        sys.exit(1)

    # Output path
    base = os.path.splitext(fpath)[0]
    output_path = f"{base}_Final.xlsx"

    # 1. Load
    print("\n[INFO] Reading pair data...")
    pairs = load_pairs(fpath)
    if not pairs:
        print("   No valid pairs — exiting.")
        return

    # 2. Normalize
    print(f"\n[INFO] Normalizing {len(pairs)} pairs...")
    m_changed = 0
    for p in pairs:
        sp = p.get("species", "unknown")
        m_norm = normalize_mirna(p["mirna_raw"], species=sp)
        g_norm = normalize_gene(p["gene_raw"], species=sp)
        if m_norm != p["mirna_raw"]:
            m_changed += 1
        p["mirna_norm"] = m_norm
        p["gene_norm"] = g_norm

    print(f"   miRNA normalized: {m_changed} changed")
    print(f"   Genes preserved as-is")

    # 3. Deduplicate
    print(f"\n[INFO] Deduplicating...")
    deduped = deduplicate(pairs)
    print(f"   Before: {len(pairs)} entries")
    print(f"   After:  {len(deduped)} groups")
    if len(pairs) > len(deduped):
        print(f"   Merged {len(pairs) - len(deduped)} duplicate pairs")

    # Multi-PMID support stats
    multi_pmid = sum(1 for d in deduped if d['pmid_count'] >= 2)
    if multi_pmid:
        print(f"   {multi_pmid} groups with >=2 PMID support")

    # 4. Resolve gene IDs (rice → ricedata.cn, others → NCBI)
    gene_map = resolve_gene_ids(deduped)
    gene_total = len(gene_map)
    gene_resolved = sum(1 for v in gene_map.values() if v.get("RAP") or v.get("NCBI"))

    # 5. Output Excel
    print(f"\n[INFO] Generating Excel...")
    export_excel(pairs, deduped, gene_map, gene_resolved, gene_total, fpath, output_path)

    # 5. Print examples for review
    print(f"\n{'='*60}")
    print(f"  Normalization examples (first 15 changes)")
    print(f"{'='*60}")
    shown = 0
    for p in pairs:
        if shown >= 15:
            break
        if p["mirna_raw"] != p["mirna_norm"]:
            print(f"  {p['mirna_raw']:25s} → {p['mirna_norm']:25s}")
            shown += 1

    if shown == 0:
        print("  (All names already in standard format — no changes needed)")

    print(f"\n{'='*60}")
    print(f"  Done!")
    print(f"  Normalized + deduplicated output: {output_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else None
    main(input_file)
