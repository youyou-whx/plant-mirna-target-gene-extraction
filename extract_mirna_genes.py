"""
miRNA–Target Gene Relationship Extraction from Literature Abstracts

Extracts miRNA–gene target pairs from batch abstract files.
Output: Excel (.xlsx), retaining only records with miRNA–gene pairs.

Columns: miRNA, Target Gene, Source Sentence, Article Title, PMID, DOI

Usage: python extract_mirna_genes.py <abstracts.txt> [output.xlsx]
"""

import re
import sys
import os
import subprocess
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ============================================================
# 0a. Blacklist — terms that are NOT gene names
# ============================================================
BLACKLIST = {
    # ----- Lab techniques / tools -----
    "PCR", "qPCR", "RT-PCR", "RT-qPCR", "qRT-PCR", "RNA-seq", "scRNA-seq",
    "ChIP", "ChIP-seq", "CLIP", "CLIP-seq", "RIP", "RIP-seq",
    "RNAi", "CRISPR", "Cas9", "TALEN", "ZFN",
    "GFP", "GUS", "LUC", "RFP", "YFP", "CFP", "mCherry",
    "GST", "MBP", "His", "FLAG", "HA", "Myc", "T7",
    # ----- Promoters / vectors -----
    "35S", "CaMV35S", "Ubi", "Actin", "NOS",
    # ----- Experimental materials / treatments -----
    "WT", "OE", "amiRNA", "MIM", "STTM",
    "Ctrl", "Control",
    # ----- Molecule types -----
    "cDNA", "mRNA", "tRNA", "rRNA", "siRNA", "sRNA", "snRNA", "snoRNA",
    "ncRNA", "lncRNA", "circRNA", "pri-miRNA", "pre-miRNA",
    "dsRNA", "ssRNA", "ssDNA", "dsDNA",
    # ----- Plant hormones -----
    "ABA", "GA", "GA3", "IAA", "JA", "SA", "BR", "ET", "ETH",
    "MeJA", "ACC", "SL", "GR24",
    # ----- Signaling molecules / metabolites -----
    "ROS", "H2O2", "NO", "NO3", "NH4", "Pi",
    "ATP", "GTP", "ADP", "GDP", "NADPH", "NADH",
    "Ca2", "Ca2+", "K+", "Na+", "Cl-",
    # ----- Statistics / units -----
    "SD", "SEM", "ANOVA", "LSD", "FDR",
    # ----- Databases / tools -----
    "NCBI", "BLAST", "GO", "KEGG", "COG", "Pfam",
    "RAP", "MSU", "TAIR", "Ensembl",
    # ----- Sequencing -----
    "SRA", "GEO", "DEG", "DEGs", "FPKM", "RPKM", "TPM", "CPM",
    # ----- Chromosomes -----
    "Chr1", "Chr2", "Chr3", "Chr4", "Chr5", "Chr6",
    "Chr7", "Chr8", "Chr9", "Chr10", "Chr11", "Chr12",
    # ----- Molecular biology basics -----
    "DNA", "RNA", "ORF", "UTR", "CDS", "SNP", "QTL", "SSR",
    "CTAB", "EDTA", "SDS", "PAGE", "TEMED", "APS",
    # ====== Added after first round ======
    # ----- False positives from compound / non-gene terms -----
    "LIKE",      # truncated from DICER-LIKE1 → LIKE1
    # ----- Viral proteins -----
    "NS3", "NS4", "NS2", "NS1",
    # ----- Rice cultivar names -----
    "IR64", "IR36", "IR8", "IR72", "IR24", "IR29",
    "ZH11", "ZH17",   # Zhonghua 11, Zhonghua 17
    "PB1",            # Pusa Basmati 1
    "TN1",            # Taichung Native 1
    "RP2068",
    "N22", "PKL",     # cultivar abbreviations
    # ----- Fungal / pathogen -----
    "AG1",  # Rhizoctonia solani AG1-IA (anastomosis group)
    # ----- Other false positives -----
    "LDLRAP1",       # human gene (appears in cross-species comparison)

    # ----- LLM auto-added (16 entries) -----
    "AUXIN",
    "BPH",
    "DOMAIN2",
    "DR",
    "EF",
    "FACTOR1",
    "III",
    "IR56",
    "IRBB5",
    "PSI",
    "PTI",
    "PXO86",
    "PXO99",
    "RBSDV",
    "TRANS",
    "UCL",
    "WRKY",
}

# ============================================================
# 0b. Whitelist — known plant gene names
# ============================================================
GENE_WHITELIST = {
    # AGO family (Argonaute)
    "AGO1", "AGO2", "AGO4", "AGO5", "AGO6", "AGO7", "AGO10",
    "AGO17", "AGO18",
    # DCL family (Dicer-like)
    "DCL1", "DCL2", "DCL3", "DCL4",
    # RDR family (RNA-dependent RNA polymerase)
    "RDR1", "RDR2", "RDR6",
    # SPL family (SQUAMOSA PROMOTER BINDING PROTEIN-LIKE)
    "SPL2", "SPL3", "SPL4", "SPL5", "SPL6", "SPL7", "SPL8",
    "SPL9", "SPL10", "SPL11", "SPL12", "SPL13", "SPL14",
    "SPL15", "SPL16", "SPL17", "SPL18",
    # AP2/EREBP family
    "AP2",
    # GRF family (Growth-Regulating Factor)
    "GRF1", "GRF2", "GRF3", "GRF4", "GRF5", "GRF6",
    "GRF7", "GRF8", "GRF9", "GRF10",
    # NAC family
    "NAC1", "NAC2", "NAC3", "NAC4", "NAC5", "NAC6",
    "NAC",             # family name (NAC transcription factors)
    # MYB family
    "MYB2", "MYB3", "MYB4", "MYB5", "MYB6",
    "MYB",             # family name
    # WRKY family
    "WRKY45", "WRKY55", "WRKY71",
    "WRKY",            # family name
    # TCP family
    "TCP21",
    "TCP",             # family name
    # Other known genes
    "PHO2", "CSD1", "CSD2", "CSD3", "CSD4",
    "TAS3",            # trans-acting siRNA locus
    "IPA1",            # Ideal Plant Architecture 1 (= OsSPL14)
    "HD3", "HD5",
    "HYL1", "SE",      # miRNA biogenesis components
    "DCL",             # sometimes used without number (family)
    "AGO",             # same as above
    "RDR",
    "SPL",
    "ARF",             # Auxin Response Factor family
    "FTL",             # FT-like
    "MFS1", "MFS2",    # SPX-MFS family (actually OsSPX-MFS1/2)
    "RFT1",            # RICE FLOWERING LOCUS T 1
    "TB1",             # Teosinte Branched 1
    "SLR1",            # SLENDER RICE 1
    "MPK3", "MPK6",    # MAP kinase
    "SNB",             # supernumerary bract
    "IDS1",            # indeterminate spikelet 1

    # ----- LLM auto-added (11 entries) -----
    "GAMYB",
    "GmPT5",
    "OsAFB2",
    "OsARF10",
    "OsGAMYBL2",
    "OsGRF4",
    "OsPHO2",
    "OsRPC53",
    "OsSOD2",
    "OsTCP21",
    "OsTIR1",
}

# ============================================================
# 1. Regex patterns
# ============================================================

# --- miRNA patterns (case-insensitive: catches miR, MIR, Mir, mir) ---
MIRNA_PATTERNS = [
    # Standard: osa-miR156a-5p, tae-miR156, ath-miR156a
    re.compile(r'\b[a-z]{3,5}-[Mm][Ii][Rr]\d+[a-z]*(?:-\d[ap])?\b'),
    # Legacy species prefix: OsmiR156, OsmiR156a, TaMIR5062-5A
    re.compile(r'\b(?:Os|Ta|Zm|At|Ath|Gm|Ptc|Ppe|Mdm|Sly|Vvi|Bna|Bra|'
               r'Cas|Csi|Mes|Nta|Stu|Hvu|Sbi|'
               r'Gh|Fv|Cm|Ma|Pv|Mt|Ca|Pt|Al|Cp|Lj)[Mm][Ii][Rr]\d+[a-z]*(?:-\d[ap])?\b'),
    # Generic: miR156, miR156a, miR156a-5p, MIR398
    re.compile(r'\b[Mm][Ii][Rr]\d+[a-z]*(?:-\d[ap])?\b'),
]

# --- Gene name patterns ---

# Tier 1: Database ID formats
GENE_TIER1 = re.compile(
    r'\b(?:LOC_Os\d{2}[gG]\d{5}(?:\.\d+)?|'
    r'Os\d{2}[gGtT]\d{7}(?:-\d{2})?)\b'
)

# Tier 2: Species prefix + uppercase + digits
GENE_TIER2 = re.compile(
    r'\b(?:Os|Ta|Zm|At|Ath|Gm|Ptc|Ppe|Mdm|Sly|Vvi|Bna|Bra|'
    r'Hvu|Sbi|Nta|Stu|Cas|Csi|Mes|'
    r'Gh|Fv|Cm|Ma|Pv|Mt|Ca|Pt|Al|Cp|Lj)\s*'
    r'[A-Z][A-Za-z]{1,5}\s*\d{1,2}\b'
)

# Tier 3: Bare uppercase + digits
GENE_TIER3 = re.compile(r'\b([A-Z]{2,6})(\d{1,2})\b')

# Tier 4: Bare uppercase gene family names (no digits) — e.g. ARF, NAC, WRKY
# Only matches whitelisted names or names with strong gene context signals
GENE_TIER4 = re.compile(r'\b([A-Z]{2,5})\b')

# Strong gene context signals — phrases immediately following a family name
STRONG_GENE_TAIL = re.compile(
    r'\b(transcription\s*factors?|genes?|family|families|proteins?|'
    r'domain\s*proteins?|box\s*proteins?|factors?|'
    r'repressors?|activators?|regulators?)\b',
    re.IGNORECASE
)
CONTEXT_SIGNALS = [
    'gene', 'genes', 'encodes', 'encode', 'encoded', 'encoding',
    'target', 'targets', 'targeted', 'targeting',
    'regulate', 'regulates', 'regulated', 'regulating', 'regulation',
    'express', 'expresses', 'expressed', 'expressing', 'expression',
    'overexpress', 'overexpresses', 'overexpressed', 'overexpression',
    'knockout', 'knock out', 'knock-out', 'knockdown', 'knock down',
    'mutant', 'mutants', 'mutation', 'mutations', 'mutated',
    'transcribe', 'transcribes', 'transcribed', 'transcription',
    'transgenic', 'transgene',
    'locus', 'loci', 'allele', 'alleles',
    'transcription factor',
    'repressor', 'activator', 'suppressor',
    'family', 'member', 'homolog', 'homologue', 'ortholog',
    'cleave', 'cleaves', 'cleaved', 'cleavage',
    'protein', 'proteins', 'encoding', 'encoded',
]

# --- miRNA–gene relation signal words ---
RELATION_SIGNALS = [
    'target', 'targets', 'targeted', 'targeting',
    'regulate', 'regulates', 'regulated', 'regulating', 'regulation',
    'cleave', 'cleaves', 'cleaved', 'cleavage',
    'repress', 'represses', 'repressed', 'repression',
    'interact', 'interacts', 'interaction',
    'recogni', 'recognizes', 'recognition',
    'bind', 'binds', 'binding', 'bound',
    'silence', 'silences', 'silenced', 'silencing',
    'suppress', 'suppresses', 'suppression',
    'mediate', 'mediates', 'mediated',
    'contribute', 'contributes', 'contributed',
    'involve', 'involves', 'involved',
    'require', 'requires', 'required',
    'modulate', 'modulates', 'modulated',
    'control', 'controls', 'controlled',
    'activate', 'activates', 'activated',
]

# ============================================================
# Species table — maps miRNA prefix → (gene prefix, common name)
# ============================================================
SPECIES_TABLE = {
    # Major cereals
    "osa": ("Os", "rice"),
    "tae": ("Ta", "wheat"),
    "zma": ("Zm", "maize"),
    "hvu": ("Hvu", "barley"),
    "sbi": ("Sbi", "sorghum"),
    # Model plants
    "ath": ("At", "Arabidopsis"),
    "aly": ("Al", "Arabidopsis lyrata"),
    # Legumes
    "gma": ("Gm", "soybean"),
    "pvu": ("Pv", "common bean"),
    "mtr": ("Mt", "alfalfa"),
    "cas": ("Cas", "chickpea"),
    "lja": ("Lj", "lotus"),
    # Solanaceae
    "sly": ("Sly", "tomato"),
    "stu": ("Stu", "potato"),
    "nta": ("Nta", "tobacco"),
    "can": ("Ca", "pepper"),
    # Fruits
    "ppe": ("Ppe", "peach"),
    "mdm": ("Mdm", "apple"),
    "vvi": ("Vvi", "grape"),
    "csi": ("Csi", "orange"),
    "cme": ("Cm", "melon"),
    "far": ("Fv", "strawberry"),
    "mac": ("Ma", "banana"),
    "cpa": ("Cp", "papaya"),
    # Fiber / oil crops
    "ghr": ("Gh", "cotton"),
    "bna": ("Bna", "rapeseed"),
    "bra": ("Bra", "turnip"),
    # Trees
    "ptc": ("Ptc", "poplar"),
    "pta": ("Pt", "pine"),
    "mes": ("Mes", "cassava"),
}

# Reverse lookup: gene prefix → miRNA prefix
_GENE_PREFIX_TO_MIR = {v[0].lower(): k for k, v in SPECIES_TABLE.items()}

# Text keyword → species (lowercase)
_SPECIES_KEYWORDS = {
    # Cereals
    "rice": "osa", "oryza": "osa",
    "wheat": "tae", "triticum": "tae",
    "maize": "zma", "corn": "zma", "zea": "zma",
    "barley": "hvu", "hordeum": "hvu",
    "sorghum": "sbi",
    # Model plants
    "arabidopsis": "ath", "thaliana": "ath",
    # Legumes
    "soybean": "gma", "glycine": "gma",
    "common bean": "pvu", "phaseolus": "pvu",
    "alfalfa": "mtr", "medicago": "mtr",
    "chickpea": "cas", "cicer": "cas",
    "lotus": "lja",
    # Solanaceae
    "tomato": "sly", "solanum lycopersicum": "sly",
    "potato": "stu", "solanum tuberosum": "stu",
    "tobacco": "nta", "nicotiana": "nta",
    "pepper": "can", "capsicum": "can",
    # Fruits
    "peach": "ppe", "prunus persica": "ppe",
    "apple": "mdm", "malus": "mdm",
    "grape": "vvi", "vitis": "vvi",
    "orange": "csi", "citrus": "csi",
    "melon": "cme", "cucumis": "cme", "cucumber": "cme",
    "strawberry": "far", "fragaria": "far",
    "banana": "mac", "musa": "mac",
    "papaya": "cpa", "carica": "cpa",
    # Fiber / oil crops
    "cotton": "ghr", "gossypium": "ghr",
    "rapeseed": "bna", "brassica napus": "bna", "canola": "bna",
    "turnip": "bra", "brassica rapa": "bra",
    # Trees
    "poplar": "ptc", "populus": "ptc",
    "pine": "pta", "pinus": "pta",
    "cassava": "mes", "manihot": "mes",
}


def detect_species(mirnas, genes, text, title=""):
    """
    Detect the dominant plant species using title-first, then weighted voting.
    Returns common name (e.g. "rice") or "unknown".
    """
    # 0. Title check — most reliable single source
    if title:
        title_lower = title.lower()
        title_hits = set()
        for keyword, mir_key in _SPECIES_KEYWORDS.items():
            if keyword in title_lower:
                title_hits.add(SPECIES_TABLE[mir_key][1])
        if len(title_hits) == 1:
            return title_hits.pop()

    votes = {}  # common_name → score

    # 1. Database IDs — weight 10 (format is species-specific, near-certain)
    for name, *_rest in mirnas + genes:
        if name.startswith('LOC_Os') or re.match(r'^Os\d{2}[gGtT]', name):
            votes["rice"] = votes.get("rice", 0) + 10

    # 2. miRNA standard prefix — weight 8 (osa-miR156 → osa → rice)
    for name, *_rest in mirnas:
        m = re.match(r'^([a-z]{3,5})-', name)
        if m and m.group(1) in SPECIES_TABLE:
            common = SPECIES_TABLE[m.group(1)][1]
            votes[common] = votes.get(common, 0) + 8

    # 3. Gene species prefix — weight 5 (OsSPL14 → Os → rice)
    # Only counts names with digits (filters bare abbreviations like ATHB, REV)
    for name, *_rest in genes:
        if not re.search(r'\d', name):
            continue  # skip digit-less names — not reliable species signals
        if len(name) >= 2:
            prefix = name[:2].lower()
            if prefix in _GENE_PREFIX_TO_MIR:
                common = SPECIES_TABLE[_GENE_PREFIX_TO_MIR[prefix]][1]
                votes[common] = votes.get(common, 0) + 5
            elif len(name) >= 3:
                prefix3 = name[:3].lower()
                if prefix3 in _GENE_PREFIX_TO_MIR:
                    common = SPECIES_TABLE[_GENE_PREFIX_TO_MIR[prefix3]][1]
                    votes[common] = votes.get(common, 0) + 5

    # 4. Text keywords — weight 1 per occurrence
    text_lower = text.lower()
    for keyword, mir_key in _SPECIES_KEYWORDS.items():
        n = text_lower.count(keyword)
        if n > 0:
            common = SPECIES_TABLE[mir_key][1]
            votes[common] = votes.get(common, 0) + n

    if votes:
        return max(votes, key=votes.get)
    return "unknown"


def extract_mirnas(text):
    """Return [(name, start, end), ...]"""
    results = []
    seen = set()
    for pattern in MIRNA_PATTERNS:
        for m in pattern.finditer(text):
            name = m.group(0)
            if name not in seen:
                # Avoid substring match: bare "miR" inside "osa-miR156"
                if name[:3].lower() == 'mir' and m.start() > 0:
                    prev_char = text[m.start() - 1]
                    if prev_char.isalpha() or prev_char == '-':
                        continue
                seen.add(name)
                results.append((name, m.start(), m.end()))
    return results


def extract_genes(text):
    """Return [(name, confidence, start, end), ...]"""
    results = []
    seen = set()

    # Tier 1: database IDs
    for m in GENE_TIER1.finditer(text):
        name = m.group(0)
        if name not in seen and name not in BLACKLIST:
            seen.add(name)
            results.append((name, "Confirmed", m.start(), m.end()))

    # Tier 2: species prefix + uppercase + digits
    for m in GENE_TIER2.finditer(text):
        name = m.group(0).replace(' ', '')
        if name not in seen and name not in BLACKLIST and not _is_generic(name):
            seen.add(name)
            results.append((name, "High confidence", m.start(), m.end()))

    # Tier 3: bare uppercase + digits
    for m in GENE_TIER3.finditer(text):
        prefix = m.group(1)
        number = m.group(2)
        name = f"{prefix}{number}"

        if name in seen:
            continue
        if name in BLACKLIST or prefix in BLACKLIST:
            continue
        if _crosses_hyphen_boundary(text, m.start()):
            continue

        # Whitelist boost: skip context check, accept directly
        if name in GENE_WHITELIST or prefix in GENE_WHITELIST:
            confidence = "High confidence"  # whitelist bumps it up one tier
        elif _has_gene_context(text, m.start(), m.end()):
            confidence = "Unverified"
        else:
            continue

        seen.add(name)
        results.append((name, confidence, m.start(), m.end()))

    # Tier 4: bare uppercase letters (no digits), e.g. ARF, NAC, WRKY
    for m in GENE_TIER4.finditer(text):
        name = m.group(1)

        if name in seen:
            continue
        if name in BLACKLIST:
            continue
        if _crosses_hyphen_boundary(text, m.start()):
            continue

        # Condition a: in whitelist → High confidence
        if name in GENE_WHITELIST:
            confidence = "High confidence"
        # Condition b: not in whitelist, but has strong gene context
        # (e.g., followed by "transcription factor") → Unverified
        elif _has_strong_gene_context(text, m.end()):
            confidence = "Unverified"
        else:
            continue

        seen.add(name)
        results.append((name, confidence, m.start(), m.end()))

    return results


def _is_generic(name):
    generics = {'OsDNA', 'OsRNA', 'OsmRNA', 'OscDNA', 'OsPCR', 'OsWT', 'OsOE', 'OsCK'}
    return name in generics


def _crosses_hyphen_boundary(text, start):
    """
    Check if a candidate crosses a hyphen boundary (false positive source).
    E.g. "LIKE1" inside "DICER-LIKE1" — because LIKE1 is preceded by a hyphen.
    """
    if start > 0 and text[start - 1] == '-':
        return True
    return False


def _has_gene_context(text, start, end):
    ctx_start = max(0, start - 120)
    ctx_end = min(len(text), end + 120)
    context = text[ctx_start:ctx_end].lower()
    return any(s.lower() in context for s in CONTEXT_SIGNALS)


def _has_strong_gene_context(text, pos):
    """
    Check whether text immediately following `pos` contains strong gene family signals.
    E.g. "ARF transcription factors", "NAC family", "WRKY genes".

    Conditions:
      1. Within 30 chars after `pos`, STRONG_GENE_TAIL pattern must match.
      2. Within 80 chars before `pos`, a gene context signal word must appear.
    """
    # 1. Immediately followed by family signal
    tail = text[pos:pos + 30]
    if not STRONG_GENE_TAIL.search(tail):
        return False

    # 2. Gene context before the candidate
    before = text[max(0, pos - 80):pos].lower()
    strong_signals = ['gene', 'genes', 'encod', 'target', 'regulat',
                      'express', 'transcri', 'protein']
    if any(s in before for s in strong_signals):
        return True

    # 3. Or preceded by "and" enumeration
    # (e.g. "...SIP19, auxin response factors (ARF) and NAC...")
    if re.search(r'\band\b\s*$', before):
        return True

    return False


# ============================================================
# Sentence splitting — chunk full text into complete sentences
# ============================================================

# Common abbreviations that do NOT end a sentence
_SENT_ABBREV = {
    'al', 'fig', 'figs', 'eq', 'eqs', 'eqn', 'ref', 'refs', 'vol', 'vols',
    'vs', 'etc', 'et', 'spp', 'sp', 'var', 'vars', 'eds', 'approx',
    'dept', 'dr', 'mr', 'mrs', 'ms', 'prof', 'inc', 'ltd', 'co', 'corp',
    'no', 'nos',
    # Month abbreviations
    'jan', 'feb', 'mar', 'apr', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
    # Cultivars
    'cv', 'cvs',
    # Latin
    'e.g', 'i.e', 'cf', 'viz',
}


def split_sentences(text):
    """
    Split full text into complete sentences using period/question/exclamation
    mark + space + uppercase letter/digit boundaries.

    Returns [(start_pos, end_pos), ...] — each sentence's position in the original text.

    Avoids false splits at abbreviations (et al., Fig. 1, cv. ZR02).
    Also treats double newlines (paragraph boundaries) as sentence boundaries.
    """
    sentences = []
    start = 0
    i = 0
    n = len(text)

    while i < n:
        if text[i] in '.!?':
            # Skip whitespace after punctuation
            j = i + 1
            while j < n and text[j] in ' \t\r':
                j += 1

            # Only treat as sentence boundary if followed by uppercase or digit
            if j < n and (text[j].isupper() or text[j].isdigit()):
                # Check if the word before punctuation is an abbreviation
                word_start = i - 1
                while word_start >= 0 and text[word_start].isalpha():
                    word_start -= 1
                word_before = text[word_start + 1:i].lower()

                if word_before not in _SENT_ABBREV:
                    sentences.append((start, j))
                    start = j
                    i = j
                    continue

        # Double newline → paragraph / section boundary
        if text[i] == '\n' and i + 1 < n and text[i + 1] == '\n':
            if start < i:
                # Trim trailing whitespace
                end = i
                while end > start and text[end - 1] in ' \t\r\n':
                    end -= 1
                if end > start:
                    sentences.append((start, end))
            start = i + 2
            i = start
            continue

        i += 1

    # Final segment
    if start < n:
        end = n
        while end > start and text[end - 1] in ' \t\r\n':
            end -= 1
        if end > start:
            sentences.append((start, end))

    return sentences


def _find_sent_index(sentences, pos):
    """Return the index of the sentence containing `pos`, or the nearest one."""
    if not sentences:
        return 0
    for idx, (s, e) in enumerate(sentences):
        if s <= pos < e:
            return idx
    if pos < sentences[0][0]:
        return 0
    return len(sentences) - 1


def get_sentence_span(text, sentences, start_pos, end_pos):
    """
    From the sentence containing `start_pos` to the sentence containing `end_pos`
    (inclusive). Returns all complete sentences concatenated as a single string.
    Ensures each sentence begins and ends at true sentence boundaries.
    """
    if not sentences:
        return get_sentence_fallback(text, start_pos, end_pos)

    si = _find_sent_index(sentences, start_pos)
    ei = _find_sent_index(sentences, end_pos)
    if si > ei:
        si, ei = ei, si

    parts = []
    for idx in range(si, ei + 1):
        s, e = sentences[idx]
        part = text[s:e].strip()
        if part:
            parts.append(part)

    return ' '.join(parts)


def get_sentence_fallback(text, start, end):
    """Fallback sentence extraction when split_sentences() fails."""
    sent_start = start
    for i in range(start, max(0, start - 800), -1):
        if text[i] in '.!?\n':
            sent_start = i + 1
            break
    else:
        sent_start = max(0, start - 800)

    sent_end = end
    for i in range(end, min(len(text), end + 800)):
        if text[i] in '.!?\n':
            sent_end = i + 1
            break
    else:
        sent_end = min(len(text), end + 800)

    return text[sent_start:sent_end].strip()


def _has_relation_context(sentence):
    """Check whether the sentence contains miRNA–target relation signal words."""
    sent_lower = sentence.lower()
    return any(sig in sent_lower for sig in RELATION_SIGNALS)


# ============================================================
# 2. File parsing
# ============================================================
def parse_abstracts_file(filepath):
    """
    Parse the abstracts file.
    Each record starts with "N. JournalName..."
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Split by "\nN. " pattern (marks the start of a new record)
    raw_blocks = re.split(r'\n(?=\d+\.\s+\w)', content)

    # Merge blocks that were incorrectly split
    # (the first block may be content before the first record, e.g. blank lines)
    blocks = []
    for block in raw_blocks:
        block = block.strip()
        if block and len(block) > 50:
            blocks.append(block)

    return blocks


def extract_metadata(block):
    """Extract PMID, DOI, and title from a single abstract block."""
    pmid = None
    doi = None
    title = None

    # PMID: usually at the end "PMID: XXXXXX"
    pmid_match = re.search(r'PMID[:：]\s*(\d+)', block, re.IGNORECASE)
    if pmid_match:
        pmid = pmid_match.group(1)

    # DOI: find across the whole block (10.xxxx/... format is unique)
    doi_match = re.search(r'\b(10\.\d{4,}/[^\s]+)', block)
    if doi_match:
        doi = doi_match.group(1).rstrip('.')

    # --- Title extraction ---
    lines = block.split('\n')

    # Step 1: Find where the journal header ends
    # Journal header = L0 + all continuation lines (starting with Epub/date/year),
    # until a blank line is encountered
    header_end = 0  # points to the last line of the journal header
    for i, line in enumerate(lines):
        if i == 0:
            continue  # L0 is always journal info
        stripped = line.strip()
        if not stripped:
            # Blank line → journal header ends
            header_end = i
            break
        # Check if this is a journal info continuation (date, Epub, or indented)
        if _is_journal_continuation(stripped):
            header_end = i
            continue
        else:
            # Not a continuation → we've passed the journal header
            header_end = i - 1
            break

    # Step 2: Extract title from after the journal header (may span multiple lines)
    title_lines = []
    for i in range(header_end + 1, len(lines)):
        stripped = lines[i].strip()
        if not stripped:
            if title_lines:
                break  # blank line → title ends
            continue
        # Author line or metadata → title ends
        if re.search(r'\(\d+\)|@', stripped):
            break
        if re.match(r'(Author info|DOI|PMID|PMCID|Conflict)', stripped, re.IGNORECASE):
            break
        # Collect title lines
        title_lines.append(stripped)

    if title_lines:
        title = ' '.join(title_lines).rstrip('.')

    return pmid, doi, title


def _is_journal_continuation(line):
    """
    Determine whether a line is a continuation of the journal info header (not the title).

    Continuation patterns: Epub prefix, bare date (Mon DD.), year-prefixed date,
    doi: prefix, indented lines.
    """
    # Epub prefix
    if re.match(r'^Epub\b', line):
        return True
    # Bare date: e.g. "Mar 7." "May 20." "2015 Jul 17."
    if re.match(r'^(?:\d{4}\s+)?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d+\.?$', line):
        return True
    # doi: prefix
    if re.match(r'^doi:', line, re.IGNORECASE):
        return True
    # Year prefix (e.g. "2024 Mar 7.")
    if re.match(r'^\d{4}\s', line) and len(line) < 20:
        return True
    # eCollection (e.g. "eCollection 2017 Oct.")
    if re.match(r'^eCollection\b', line, re.IGNORECASE):
        return True
    # DOI continuation (e.g. "10.1002/anie.202214987. Epub 2023 Feb 14.")
    if re.match(r'^10\.\d{4,}/', line):
        return True
    # Online ahead of print
    if re.match(r'^Online\s+ahead\s+of\s+print', line, re.IGNORECASE):
        return True
    # Wrapped continuation (starts lowercase, short — e.g., "print." from "Online ahead of print.")
    if line[0].islower() and len(line) < 40:
        return True
    return False


def get_abstract_body(block):
    """Extract the abstract body text (excluding metadata lines)."""
    lines = block.split('\n')
    body_lines = []
    in_body = False
    for line in lines:
        stripped = line.strip()
        # Skip metadata
        if re.match(r'(DOI|PMID|PMCID|Conflict|Author info)', stripped, re.IGNORECASE):
            continue
        if re.search(r'\(\d+\).*@|^\d+\.\s', stripped):
            continue
        if not stripped:
            if in_body:
                break  # blank line after body has started → end
            continue
        # Skip author lines
        if re.search(r'\(\d+\)|^[A-Z][a-z]+ [A-Z]\.?$', stripped):
            continue
        body_lines.append(stripped)
        in_body = True

    return ' '.join(body_lines)


# ============================================================
# 3. Excel output (miRNA–gene pairs only)
# ============================================================

HEADER_FONT = Font(name='Consolas', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

HIGH_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
MEDIUM_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
NORMAL_FONT = Font(name='Consolas', size=10)
WRAP_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def build_excel(pairs, output_path):
    """
    Sheet 1: miRNA–Target Gene Pairs (detail)
    Sheet 2: Statistics (overview)
    """
    wb = openpyxl.Workbook()

    # ========================
    # Sheet 1: Pair detail
    # ========================
    ws1 = wb.active
    ws1.title = "miRNA-Target Gene Pairs"

    headers = ["#", "miRNA", "Target Gene", "Relation Confidence",
               "Relation Type", "Source Sentence", "Article Title", "PMID", "DOI",
               "Species"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for i, pair in enumerate(pairs):
        row = i + 2
        values = [
            i + 1,
            pair["mirna"],
            pair["gene"],
            pair["confidence"],
            pair["relation_type"],
            pair["sentence"],
            pair["title"] or "",
            pair["pmid"] or "",
            pair["doi"] or "",
            pair.get("species", "unknown"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT if col in (6, 7) else CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            # Color-code relation confidence
            if col in (2, 3, 4) and pair["confidence"].startswith(("Confirmed", "High confidence")):
                cell.fill = HIGH_FILL
            elif col in (2, 3, 4) and "Unverified" in pair["confidence"]:
                cell.fill = MEDIUM_FILL

    # Column widths
    col_widths = [6, 14, 14, 18, 14, 65, 45, 12, 28, 14]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(pairs)+1}"

    # ========================
    # Sheet 2: Statistics
    # ========================
    ws2 = wb.create_sheet(title="Statistics")

    # Dedup stats
    unique_pmids = set(p["pmid"] for p in pairs if p["pmid"])
    unique_mirnas = set(p["mirna"] for p in pairs)
    unique_genes = set(p["gene"] for p in pairs)

    stats = [
        ("Total miRNA–target gene pairs", len(pairs)),
        ("Unique PMIDs", len(unique_pmids)),
        ("Unique miRNAs", len(unique_mirnas)),
        ("Unique genes", len(unique_genes)),
    ]

    title_cell = ws2.cell(row=1, column=1, value="miRNA–Target Gene Extraction Statistics")
    title_cell.font = Font(name='Consolas', bold=True, size=14, color='2F5496')
    ws2.merge_cells('A1:B1')

    for i, (label, value) in enumerate(stats):
        row = i + 3
        ws2.cell(row=row, column=1, value=label).font = Font(name='Consolas', bold=True, size=11)
        ws2.cell(row=row, column=2, value=value).font = NORMAL_FONT

    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 20

    wb.save(output_path)
    wb.close()
    print(f"   Excel saved to: {output_path}")


# ============================================================
# 4. Pairing logic
# ============================================================
def pair_mirna_gene(mirnas, genes, text, abstract_body_text):
    """
    Pair miRNAs with genes: co-occurrence in the same sentence → one pair record.

    Cross-match detection:
    - In a crowded sentence (multiple miRNAs + multiple genes), only pairs with
      explicit connections (miRNA/gene or miRNA-gene) are marked High confidence.
    - Other same-sentence pairs are downgraded to "associated" for LLM review.

    Returns: [{mirna, gene, gene_confidence, relation_type, sentence}, ...]
    """
    pairs = []

    # Pre-split full text into sentences
    all_sentences = split_sentences(text)

    # Pre-compute which sentence each miRNA / gene falls in
    mirna_si = {}
    for m_name, m_start, m_end in mirnas:
        mirna_si[m_name] = _find_sent_index(all_sentences, m_start)

    gene_si = {}
    for g_name, g_conf, g_start, g_end in genes:
        gene_si[g_name] = _find_sent_index(all_sentences, g_start)

    for m_name, m_start, m_end in mirnas:
        m_si = mirna_si[m_name]

        for g_name, g_conf, g_start, g_end in genes:
            g_si = gene_si[g_name]

            # ── Proximity check (sentence-based, not character-based) ──
            # Same sentence: always pair, regardless of character distance
            # Adjacent sentences: pair only if char distance ≤ 400 (safety valve)
            # Gap ≥ 2 sentences: skip
            sent_dist = abs(m_si - g_si)
            if sent_dist == 0:
                pass  # same sentence → always pair
            elif sent_dist == 1 and abs(m_start - g_start) <= 400:
                pass  # adjacent sentences + not too far
            else:
                continue

            # ── Pick the single best sentence that shows the relationship ──
            sentence = _pick_best_sentence(
                text, all_sentences, m_si, g_si, m_start, m_end, g_start, g_end
            )

            # ── Cross-match detection ──
            si_min = min(m_si, g_si)
            si_max = max(m_si, g_si)

            mirnas_in_span = set()
            genes_in_span = set()
            for nm, ms, me in mirnas:
                if si_min <= mirna_si.get(nm, -1) <= si_max:
                    mirnas_in_span.add(nm)
            for gn, gc, gs, ge in genes:
                if si_min <= gene_si.get(gn, -1) <= si_max:
                    genes_in_span.add(gn)

            is_crowded = len(mirnas_in_span) > 1 or len(genes_in_span) > 1
            is_explicit = _has_explicit_link(text, min(m_start, g_start),
                                             max(m_end, g_end), m_name, g_name)

            # ── Relation confidence (how certain is the targeting relationship?) ──
            rel_type = _classify_relation(sentence)

            if is_crowded and not is_explicit:
                # Multiple entities crowded + no explicit link → uncertain
                rel_conf = "Unverified (crowded)"
                rel_type = "associated"
            elif is_crowded and is_explicit:
                # Crowded but has explicit link (e.g. miR396b/GRF6) → High confidence
                rel_conf = "High confidence"
            elif is_explicit:
                # Non-crowded with explicit link → High confidence
                # (slash/hyphen itself implies targeting)
                rel_conf = "High confidence"
            elif rel_type in ("cleavage", "repression", "targeting"):
                # Non-crowded + strong targeting verb → Confirmed
                rel_conf = "Confirmed"
            elif rel_type in ("regulation", "binding", "interaction"):
                # Non-crowded + regulatory/binding verb → High confidence
                rel_conf = "High confidence"
            else:
                # Non-crowded but no targeting verb → Unverified
                rel_conf = "Unverified"

            # ── Combined confidence (output to Excel) ──
            # Takes the worse of gene confidence and relation confidence
            if g_conf == "Unverified" and rel_conf.startswith("Unverified"):
                final_conf = "Unverified"
            elif g_conf == "Unverified":
                final_conf = f"{rel_conf} (gene unverified)"
            elif rel_conf.startswith("Unverified"):
                final_conf = rel_conf  # gene OK, relation uncertain
            else:
                final_conf = rel_conf

            pairs.append({
                "mirna": m_name,
                "gene": g_name,
                "gene_confidence": g_conf,          # internal: is the gene name reliable?
                "relation_confidence": rel_conf,    # is the targeting relationship certain?
                "confidence": final_conf,            # displayed in Excel column
                "relation_type": rel_type,
                "sentence": sentence,
            })

    return pairs


def _pick_best_sentence(text, sentences, m_si, g_si, m_start, m_end, g_start, g_end):
    """
    From the span of sentences between miRNA and gene, pick the single sentence
    that best demonstrates the relationship.

    Same sentence → return it directly.
    Different sentences → score each by number of relation signal word hits,
    return the highest-scoring sentence.
    """
    # Same sentence: return directly
    if m_si == g_si:
        s, e = sentences[m_si]
        return text[s:e].strip()

    # Different sentences: score and pick best
    best_sent = ""
    best_score = -1
    for si in range(min(m_si, g_si), max(m_si, g_si) + 1):
        s, e = sentences[si]
        sent_text = text[s:e].strip()
        if not sent_text:
            continue

        # +1 point per relation signal word hit
        score = 0
        sent_lower = sent_text.lower()
        for sig in RELATION_SIGNALS:
            if sig in sent_lower:
                score += 1

        if score > best_score:
            best_score = score
            best_sent = sent_text

    # Fallback: if no relation signal words found anywhere, use the gene's sentence
    if best_score == 0 and best_sent == "":
        s, e = sentences[g_si]
        return text[s:e].strip()

    return best_sent if best_sent else text[sentences[g_si][0]:sentences[g_si][1]].strip()


def _has_explicit_link(text, sent_start, sent_end, mirna, gene):
    """Check whether miRNA and gene have an explicit / or - link in the sentence."""
    sentence = text[sent_start:sent_end]
    # miRNA/gene or gene/miRNA
    if f"{mirna}/{gene}" in sentence or f"{gene}/{mirna}" in sentence:
        return True
    # miRNA-gene (careful: LOC_Os01g01700 already has hyphens)
    if f"{mirna}-{gene}" in sentence:
        return True
    return False


def _classify_relation(sentence):
    """Classify the relationship type based on verbs in the sentence."""
    sent = sentence.lower()
    if any(w in sent for w in ['cleave', 'cleaves', 'cleaved', 'cleavage']):
        return "cleavage"
    if any(w in sent for w in ['repress', 'represses', 'repressed', 'repression',
                                 'silence', 'silences', 'silenced', 'silencing',
                                 'suppress', 'suppresses', 'suppression']):
        return "repression"
    if any(w in sent for w in ['target', 'targets', 'targeted', 'targeting']):
        return "targeting"
    if any(w in sent for w in ['regulate', 'regulates', 'regulated', 'regulation',
                                 'mediate', 'mediates', 'mediated',
                                 'contribute', 'contributes', 'contributed',
                                 'modulate', 'modulates', 'modulated',
                                 'control', 'controls', 'controlled',
                                 'activate', 'activates', 'activated',
                                 'involve', 'involves', 'involved',
                                 'require', 'requires', 'required']):
        return "regulation"
    if any(w in sent for w in ['bind', 'binds', 'binding', 'bound',
                                 'recogni', 'recognizes', 'recognition']):
        return "binding"
    if any(w in sent for w in ['interact', 'interacts', 'interaction']):
        return "interaction"
    return "associated"


# ============================================================
# 5. Main
# ============================================================
def main(input_file, output_file=None):
    print(f"Reading file: {input_file}")
    blocks = parse_abstracts_file(input_file)
    print(f"Found {len(blocks)} abstracts\n")

    if output_file is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_file = os.path.join(script_dir, "miRNA-Target_Gene_Pairs.xlsx")

    all_pairs = []
    stats = {
        "total": len(blocks),
        "has_mirna": 0,
        "has_gene": 0,
        "has_both": 0,
        "has_pair": 0,
        "total_mirna": 0,
        "total_gene": 0,
    }

    for i, block in enumerate(blocks):
        pmid, doi, title = extract_metadata(block)
        body = get_abstract_body(block)
        full_text = block

        mirnas = extract_mirnas(full_text)
        genes = extract_genes(full_text)

        if mirnas:
            stats["has_mirna"] += 1
            stats["total_mirna"] += len(mirnas)
        if genes:
            stats["has_gene"] += 1
            stats["total_gene"] += len(genes)
        if mirnas and genes:
            stats["has_both"] += 1

        # Species detection
        species = detect_species(mirnas, [(g[0], g[1], g[2], g[3]) for g in genes], full_text, title or "")

        # Pairing
        if mirnas and genes:
            pairs = pair_mirna_gene(mirnas, genes, full_text, body)
            if pairs:
                stats["has_pair"] += 1
                for p in pairs:
                    p["pmid"] = pmid
                    p["doi"] = doi
                    p["title"] = title
                    p["abstract_index"] = i + 1
                    p["species"] = species
                all_pairs.extend(pairs)

        if (i + 1) % 200 == 0:
            print(f"   Processed {i+1}/{len(blocks)} ... found {len(all_pairs)} pairs")

    # Generate Excel
    print(f"\nGenerating Excel...")
    build_excel(all_pairs, output_file)

    # Terminal summary
    print(f"\n{'='*55}")
    print(f"  Done!")
    print(f"{'='*55}")
    print(f"  Total abstracts:              {stats['total']}")
    print(f"  With miRNA:                   {stats['has_mirna']}")
    print(f"  With gene:                    {stats['has_gene']}")
    print(f"  With both miRNA+gene:         {stats['has_both']}")
    print(f"  With paired relationships:    {stats['has_pair']}")
    print(f"  {'─'*40}")
    print(f"  Total miRNA–target gene pairs: {len(all_pairs)}")
    if all_pairs:
        unique_m = len(set(p["mirna"] for p in all_pairs))
        unique_g = len(set(p["gene"] for p in all_pairs))
        print(f"  Unique miRNAs:                 {unique_m}")
        print(f"  Unique genes:                  {unique_g}")
    print(f"{'='*55}")
    print(f"  Excel file: {output_file}")
    print(f"{'='*55}")

    if '--auto-review' in sys.argv:
        print(f"\n[INFO] Auto-review enabled — launching LLM review...\n")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        llm_review_script = os.path.join(script_dir, "llm_review.py")
        result = subprocess.run([sys.executable, llm_review_script], cwd=script_dir)
        if result.returncode != 0:
            print(f"\n[WARN] LLM review exited with code {result.returncode}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_mirna_genes.py <abstracts.txt> [output.xlsx]")
        print("Example: python extract_mirna_genes.py abstract-ricemirna-set.txt")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    main(input_path, output_path)
