"""
LLM Review Script

Input:  abstract-ricemirna-set_miRNA-Target_Gene_Pairs.xlsx (regex extraction output)
Output: abstract-ricemirna-set_miRNA-Target_Gene_Pairs_LLMReviewed.xlsx

Cost estimate: ~200 unverified entries ≈ 0.1 CNY
"""

import json, os, sys, time
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Configuration
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_XLSX = os.path.join(SCRIPT_DIR, "miRNA-Target_Gene_Pairs.xlsx")
OUTPUT_XLSX = os.path.join(SCRIPT_DIR, "miRNA-Target_Gene_Pairs_Reviewed.xlsx")
PROGRESS_FILE = os.path.join(SCRIPT_DIR, "llm_review_progress.json")  # checkpoint / resume

# LLM configuration
LLM_API_URL = "https://api.deepseek.com/v1/chat/completions"
LLM_MODEL = "deepseek-chat"
LLM_MAX_RETRIES = 3
LLM_SLEEP_BETWEEN = 0.1       # seconds between batches (smaller is fine with batching)
BATCH_SAVE_EVERY = 100         # save progress every N entries
ENTRIES_PER_BATCH = 5          # entries per API call (core speed optimization)

# ============================================================
# Styles
# ============================================================
HEADER_FONT = Font(name='Consolas', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

NORMAL_FONT = Font(name='Consolas', size=10)
BOLD_FONT = Font(name='Consolas', size=10, bold=True)
WRAP_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


# ============================================================
# 1. Load Excel, pick out entries needing LLM review
# ============================================================
def load_uncertain_entries():
    """Return [(excel_row, entry_dict), ...]"""
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb['miRNA-Target Gene Pairs']

    entries = []
    for row in range(2, ws.max_row + 1):
        conf = str(ws.cell(row=row, column=4).value or '')
        rel = str(ws.cell(row=row, column=5).value or '')

        # Send to LLM if: relation is uncertain ("Unverified") or targeting evidence is weak ("associated")
        if 'unverified' in conf.lower() or rel == 'associated':
            entry = {
                'idx': row - 1,
                'excel_row': row,
                'mirna': str(ws.cell(row=row, column=2).value or ''),
                'gene': str(ws.cell(row=row, column=3).value or ''),
                'conf': conf,
                'rel': rel,
                'sent': str(ws.cell(row=row, column=6).value or ''),
                'title': str(ws.cell(row=row, column=7).value or ''),
                'pmid': str(ws.cell(row=row, column=8).value or ''),
                'species': str(ws.cell(row=row, column=10).value or ''),
            }
            entries.append(entry)

    wb.close()
    return entries


# ============================================================
# 2. Build prompts, call LLM
# ============================================================
def build_prompt(entry):
    """Build a prompt for a single entry (fallback mode)."""
    return f"""你是一个植物分子生物学文献审稿助手。请分析以下句子中 miRNA 与基因的关系。

miRNA: {entry['mirna']}
基因: {entry['gene']}
物种: {entry.get('species', 'unknown')}
论文标题: {entry['title']}
句子: "{entry['sent']}"

【判断任务】

1. "{entry['gene']}" 是一个真实的基因名吗？
   你可以结合你关于植物分子生物学的知识来判断。
   否 = 它是植物品种名（如IR56/IRBB5/ZH11）、菌株名（如PXO86/Guy11）、截断的词语片段（如FACTOR1来自"PROLIFERATING CELL FACTOR1"）、或人类/动物基因名
   是 = 它是一个植物基因名

2. {entry['mirna']} 是否靶向 {entry['gene']}？
   （这一步主要看句子文字，可结合你的背景知识辅助判断。）

   以下情况算"是"：
   - 出现靶向动词（targets/cleaves/regulates/suppresses/silences...）且连接这对 miRNA 和基因
   - miRNA和基因用 "/" 或 "-" 直接连写（如 miR396b/GRF6 module）
   - 明确描述为一对调控单元（如 "X/Y regulatory unit"、"X-mediated regulation of Y"）
   - 描述了切割验证实验（如 RLM-5'RACE、degradome sequencing confirmed）

3. 如果不存在靶向关系（is_target=false），判断 miRNA 和基因在这个句子里是否有**生物学关联**？
   - "related" = 句子表明两者可能为靶向关系，但你不确定，如：
     * miRNA 和基因之间有间接调控关系（通过其他因子介导）（这时请再次结合你的知识辅助判断，如果能确定有靶向关系，则归为是靶向）
     * 虽然没靶向动词，但明确暗示 miRNA 影响该基因（这时请再次结合你的知识辅助判断，如果能确定有靶向关系，则归为是靶向）
   - "unrelated" = 明显不是靶向作用关系或文章显然没有讨论证明两者间的靶向关系，如：
     * 背景罗列已知 miRNA/基因（"previous studies identified..."）
     * 引用前人研究但这句不讨论它们的关系
     * 两个名字刚好在同一句但毫无功能关联
     * 同通路、共表达、共调控（"both were upregulated"、"co-expressed"）
     * 同一生物学过程中协同作用

4. 如果存在靶向关系，类型是什么？（优先根据原文，而后根据你的背景知识辅助判断）
   cleavage = 切割/剪切
   repression = 负调控/抑制/沉默
   targeting = 靶向（有 target 关键词）
   regulation = 调控（regulate/mediate 等）

【输出格式 —— 纯JSON，不要任何其他文字】
{{"is_gene": true或false, "is_target": true或false, "relevance": "targeting/related/unrelated", "relation_type": "cleavage/repression/targeting/regulation/无", "evidence_from_text": "理由"}}
其中 evidence_from_text 必须说明判断依据：
- 若 is_gene=false：说明它实际是什么（如"PTI是免疫过程非基因"、"IR56是水稻品种名"）
- 若 is_target 相关：摘出原文关键短语（5-20字），证明靶向关系是否存在
- 若 is_target=false 且 relevance=unrelated：简述为什么无关（如"纯背景罗列，未讨论两者关系"）"""


def build_batch_prompt(entries):
    """Pack multiple entries into a single batch prompt."""
    parts = [
        "你是植物分子生物学文献审稿助手。请逐条分析以下 miRNA-基因配对关系。",
        "对每条给出独立判断，输出一个 JSON 数组。\n",
    ]
    for i, e in enumerate(entries):
        parts.append(
            f"--- [{i+1}/{len(entries)}] ---\n"
            f"miRNA: {e['mirna']}\n"
            f"基因: {e['gene']}\n"
            f"物种: {e.get('species', 'unknown')}\n"
            f"句子: \"{e['sent']}\"\n"
        )
    parts.append(
        "\n【对每条输出】\n"
        "1. is_gene: \"{gene}\" 是真实基因名吗？否=品种名/菌株/过程缩写/技术/截断词/人类基因\n"
        "2. is_target: 句子是否描述靶向关系？（targets/cleaves/regulates/斜杠连写/RLM-5'RACE等）\n"
        "3. relevance: targeting=靶向, related=可能相关但不确认, unrelated=纯共现/背景罗列\n"
        "4. relation_type: cleavage/repression/targeting/regulation/无\n"
        "5. evidence_from_text: 支持你判断的关键依据（5-20字）\n"
        "\n【输出纯JSON数组，不要其他文字】\n"
        '[{"id":1,"is_gene":true,"is_target":false,"relevance":"unrelated","relation_type":"无","evidence_from_text":"简短理由"},...]'
    )
    return '\n'.join(parts)


def call_llm_batch(prompt, api_key):
    """
    Call DeepSeek API for batch judgment.
    Returns [{is_gene, is_target, ...}, ...] list, or None on failure.
    """
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 2500,
    }

    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = requests.post(LLM_API_URL, headers=headers, json=payload, timeout=60)

            if resp.status_code == 200:
                data = resp.json()
                content = data['choices'][0]['message']['content'].strip()

                if '```' in content:
                    content = content.split('```')[1]
                    if content.startswith('json'):
                        content = content[4:]
                    content = content.strip()

                results = json.loads(content)
                # Ensure we have a list
                if isinstance(results, dict) and 'results' in results:
                    results = results['results']
                if not isinstance(results, list):
                    results = [results]

                parsed = []
                for r in results:
                    relevance = r.get('relevance', '')
                    if not relevance:
                        if r.get('is_target', False):
                            relevance = 'targeting'
                        else:
                            relevance = 'unrelated'
                    parsed.append({
                        'is_gene': r.get('is_gene', True),
                        'is_target': r.get('is_target', False),
                        'relevance': relevance,
                        'relation_type': r.get('relation_type', 'none'),
                        'explanation': r.get('evidence_from_text', r.get('explanation', '')),
                    })
                return parsed

            elif resp.status_code == 429:
                wait = (attempt + 1) * 5
                print(f"   [WARN] Rate limited (429), retrying in {wait}s...")
                time.sleep(wait)
                continue

            elif resp.status_code == 401:
                print(f"   [ERROR] Invalid API Key (401)")
                return None

            else:
                print(f"   [WARN] HTTP {resp.status_code}: {resp.text[:100]}")
                time.sleep(2)
                continue

        except requests.exceptions.Timeout:
            print(f"   [WARN] Request timeout, retry {attempt+1}/{LLM_MAX_RETRIES}")
            time.sleep(3)
        except json.JSONDecodeError as e:
            print(f"   [WARN] JSON parse error: {e}")
            time.sleep(1)
        except Exception as e:
            print(f"   [WARN] Unexpected error: {e}")
            time.sleep(1)

    return None


def call_llm(prompt, api_key):
    """
    Call DeepSeek API for a single entry (fallback when batch fails).
    Returns parsed dict, or None on failure.
    """
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,   # low temperature = more deterministic
        "max_tokens": 300,
    }

    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = requests.post(LLM_API_URL, headers=headers, json=payload, timeout=30)

            if resp.status_code == 200:
                data = resp.json()
                content = data['choices'][0]['message']['content'].strip()

                if '```' in content:
                    content = content.split('```')[1]
                    if content.startswith('json'):
                        content = content[4:]
                    content = content.strip()

                result = json.loads(content)
                # Backward compatibility: infer relevance from is_target if missing
                relevance = result.get('relevance', '')
                if not relevance:
                    if result.get('is_target', False):
                        relevance = 'targeting'
                    else:
                        relevance = 'unrelated'
                return {
                    'is_gene': result.get('is_gene', True),
                    'is_target': result.get('is_target', False),
                    'relevance': relevance,
                    'relation_type': result.get('relation_type', 'none'),
                    'explanation': result.get('evidence_from_text', result.get('explanation', '')),
                }

            elif resp.status_code == 429:
                wait = (attempt + 1) * 3
                print(f"   [WARN] Rate limited (429), retrying in {wait}s...")
                time.sleep(wait)
                continue

            elif resp.status_code == 401:
                print(f"   [ERROR] Invalid API Key (401). Check DEEPSEEK_API_KEY.")
                return None

            else:
                print(f"   [WARN] HTTP {resp.status_code}: {resp.text[:100]}")
                time.sleep(1)
                continue

        except requests.exceptions.Timeout:
            print(f"   [WARN] Request timeout, retry {attempt+1}/{LLM_MAX_RETRIES}")
            time.sleep(2)
        except json.JSONDecodeError as e:
            print(f"   [WARN] JSON parse error: {e}, LLM returned: {content[:100]}")
            time.sleep(1)
        except Exception as e:
            print(f"   [WARN] Unexpected error: {e}")
            time.sleep(1)

    return None


# ============================================================
# 3. Merge results into Excel
# ============================================================
def merge_results(entries, reviewed):
    """
    entries: all rows from original Excel
    reviewed: LLM review results {idx: {is_gene, is_target, ...}}
    """
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb['miRNA-Target Gene Pairs']

    # Add review columns (after existing 10 columns including Species)
    col_review = 11
    col_note = 12
    col_keep = 13
    ws.cell(row=1, column=col_review, value="LLM Review Result").font = HEADER_FONT
    ws.cell(row=1, column=col_review).fill = HEADER_FILL
    ws.cell(row=1, column=col_review).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=col_review).border = THIN_BORDER

    ws.cell(row=1, column=col_note, value="LLM Rationale").font = HEADER_FONT
    ws.cell(row=1, column=col_note).fill = HEADER_FILL
    ws.cell(row=1, column=col_note).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=col_note).border = THIN_BORDER

    ws.cell(row=1, column=col_keep, value="Manual Keep").font = HEADER_FONT
    ws.cell(row=1, column=col_keep).fill = HEADER_FILL
    ws.cell(row=1, column=col_keep).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=col_keep).border = THIN_BORDER

    stats = {"Confirmed": 0, "Weak association": 0,
             "Excluded (non-target)": 0, "Excluded (non-gene)": 0,
             "Not reviewed": 0}

    for row in range(2, ws.max_row + 1):
        idx = row - 1
        conf = str(ws.cell(row=row, column=4).value or '')
        rel = str(ws.cell(row=row, column=5).value or '')

        if idx in reviewed:
            rv = reviewed[idx]
            is_gene = rv['is_gene']
            is_target = rv['is_target']
            relevance = rv.get('relevance', 'unrelated')
            note = rv['explanation']

            if not is_gene:
                status, fill = "Excluded (non-gene)", RED_FILL
                stats["Excluded (non-gene)"] += 1
            elif is_target or relevance == 'targeting':
                status, fill = f"Confirmed ({rv['relation_type']})", GREEN_FILL
                stats["Confirmed"] += 1
            elif relevance == 'related':
                status, fill = "Weak association", YELLOW_FILL
                stats["Weak association"] += 1
            else:  # unrelated
                status, fill = "Excluded (non-target)", RED_FILL
                note = note or "Pure co-occurrence / enumeration, no biological relevance"
                stats["Excluded (non-target)"] += 1
        else:
            # Entries not sent for LLM review (already high confidence from regex)
            if conf in ("Confirmed", "High confidence") and rel not in ("associated",):
                status, fill = "Confirmed (regex high confidence)", GREEN_FILL
                note = "Regex high confidence, not sent for LLM review"
                stats["Confirmed"] += 1
            else:
                status, fill = "Not reviewed", YELLOW_FILL
                note = "Not reviewed by LLM (confidence value not recognized)"
                stats["Not reviewed"] += 1

        cell_s = ws.cell(row=row, column=col_review, value=status)
        cell_s.font = NORMAL_FONT
        cell_s.alignment = CENTER_ALIGNMENT
        cell_s.border = THIN_BORDER
        cell_s.fill = fill

        # Synchronize miRNA and gene column colors; clear old relation-confidence coloring
        for col in (2, 3):
            c = ws.cell(row=row, column=col)
            c.fill = fill
        ws.cell(row=row, column=4).fill = PatternFill(fill_type=None)

        cell_n = ws.cell(row=row, column=col_note, value=note)
        cell_n.font = NORMAL_FONT
        cell_n.alignment = WRAP_ALIGNMENT
        cell_n.border = THIN_BORDER

        # Manual Keep column — pre-fill "Yes" for Confirmed, empty for others
        if status.startswith("Confirmed"):
            keep_val = "Yes"
        else:
            keep_val = ""
        cell_k = ws.cell(row=row, column=col_keep, value=keep_val)
        cell_k.font = NORMAL_FONT
        cell_k.alignment = CENTER_ALIGNMENT
        cell_k.border = THIN_BORDER
        if "Weak association" in status:
            cell_k.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws.column_dimensions['K'].width = 28
    ws.column_dimensions['L'].width = 50
    ws.column_dimensions['M'].width = 14
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    # Update Statistics sheet
    ws2 = wb['Statistics']
    add_row = 10
    ws2.cell(row=add_row, column=1, value="").font = NORMAL_FONT
    ws2.cell(row=add_row+0, column=1, value="--- After LLM Review ---").font = Font(name='Consolas', bold=True, size=12, color='2F5496')
    ws2.cell(row=add_row+1, column=1, value="Final retained pairs").font = BOLD_FONT
    ws2.cell(row=add_row+1, column=2, value=stats['Confirmed'] + stats['Weak association']).font = NORMAL_FONT
    ws2.cell(row=add_row+2, column=1, value="  Confirmed").font = NORMAL_FONT
    ws2.cell(row=add_row+2, column=2, value=stats['Confirmed']).font = NORMAL_FONT
    ws2.cell(row=add_row+3, column=1, value="  Weak association (retained)").font = NORMAL_FONT
    ws2.cell(row=add_row+3, column=2, value=stats['Weak association']).font = NORMAL_FONT
    ws2.cell(row=add_row+4, column=1, value="  Excluded (non-target)").font = NORMAL_FONT
    ws2.cell(row=add_row+4, column=2, value=stats['Excluded (non-target)']).font = NORMAL_FONT
    ws2.cell(row=add_row+5, column=1, value="  Excluded (non-gene)").font = NORMAL_FONT
    ws2.cell(row=add_row+5, column=2, value=stats['Excluded (non-gene)']).font = NORMAL_FONT
    ws2.cell(row=add_row+6, column=1, value="  Total excluded").font = BOLD_FONT
    ws2.cell(row=add_row+6, column=2, value=stats['Excluded (non-target)'] + stats['Excluded (non-gene)']).font = BOLD_FONT

    wb.save(OUTPUT_XLSX)
    wb.close()
    return stats


# ============================================================
# 4. Main
# ============================================================
def main():
    # Check API Key
    api_key = os.environ.get('DEEPSEEK_API_KEY', '')
    if not api_key:
        print("=" * 60)
        print("  ERROR: DEEPSEEK_API_KEY environment variable not set")
        print("")
        print("  Register at https://platform.deepseek.com to get an API Key,")
        print("  then run:")
        print('    $env:DEEPSEEK_API_KEY = "sk-xxxxxxxx"')
        print("")
        print("  (If using a different LLM provider, update LLM_API_URL at the top of this script)")
        print("=" * 60)
        sys.exit(1)

    # 1. Load entries needing review
    print("[INFO] Reading regex extraction results...")
    entries = load_uncertain_entries()
    print(f"   {len(entries)} entries need LLM review")

    if not entries:
        print("   No uncertain entries — no LLM calls needed.")
        return

    # 2. Check for previous interrupted progress
    reviewed = {}  # {idx: LLM result}
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            reviewed = json.load(f)
            # Convert keys back to int
            reviewed = {int(k): v for k, v in reviewed.items()}
        print(f"   [INFO] Resuming from checkpoint: {len(reviewed)} entries already reviewed")

    pending = [e for e in entries if e['idx'] not in reviewed]
    print(f"   Remaining: {len(pending)} entries")

    if not pending:
        print("   All entries have been reviewed!")
    else:
        # 3. Batch LLM calls
        total_batches = (len(pending) + ENTRIES_PER_BATCH - 1) // ENTRIES_PER_BATCH
        print(f"\n[INFO] Starting batch DeepSeek API calls...")
        print(f"   Model: {LLM_MODEL}  |  {ENTRIES_PER_BATCH} entries/batch  |  {total_batches} batches total")
        print(f"   Estimated time: ~{total_batches * 3 / 60:.1f} min\n")

        success = 0
        fail = 0

        for batch_idx in range(0, len(pending), ENTRIES_PER_BATCH):
            batch = pending[batch_idx:batch_idx + ENTRIES_PER_BATCH]
            batch_no = batch_idx // ENTRIES_PER_BATCH + 1
            b_start = batch_idx + 1
            b_end = min(batch_idx + ENTRIES_PER_BATCH, len(pending))
            print(f"  [Batch {batch_no}/{total_batches}] entries {b_start}-{b_end} ", end="", flush=True)

            prompt = build_batch_prompt(batch)
            results = call_llm_batch(prompt, api_key)

            if results and len(results) == len(batch):
                for j, (entry, result) in enumerate(zip(batch, results)):
                    reviewed[entry['idx']] = result
                    success += 1
                    i = batch_idx + j
                    if result['is_gene'] and result['is_target']:
                        print(f"\n    [{i+1}] {entry['mirna']} -> {entry['gene']}  [OK] {result['relation_type']}", end="")
                    elif result['is_gene'] and result.get('relevance') == 'related':
                        print(f"\n    [{i+1}] {entry['mirna']} -> {entry['gene']}  [WEAK] weak association", end="")
                    elif result['is_gene']:
                        print(f"\n    [{i+1}] {entry['mirna']} -> {entry['gene']}  [EXCL] non-target", end="")
                    else:
                        print(f"\n    [{i+1}] {entry['mirna']} -> {entry['gene']}  [EXCL] non-gene", end="")
                print()  # newline after batch
            else:
                # Batch failed → fall back to single-entry calls
                print("[FALLBACK] Batch failed, retrying one-by-one...")
                for entry in batch:
                    result = call_llm(build_prompt(entry), api_key)
                    if result:
                        reviewed[entry['idx']] = result
                        success += 1
                    else:
                        fail += 1
                    time.sleep(0.3)

            # Periodic save
            if batch_idx + ENTRIES_PER_BATCH >= len(pending) or \
               (batch_idx // ENTRIES_PER_BATCH + 1) % (BATCH_SAVE_EVERY // ENTRIES_PER_BATCH) == 0:
                with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(reviewed, f, ensure_ascii=False, indent=2)
                print(f"      [SAVE] Progress saved ({len(reviewed)} entries)")

            time.sleep(LLM_SLEEP_BETWEEN)

        # Final save
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(reviewed, f, ensure_ascii=False, indent=2)

        print(f"\n   LLM review complete! Success: {success}, Failed: {fail}")
        if fail > 0:
            print(f"   Re-run the script to retry failed entries (checkpoint resume)")

    # 4. Merge into Excel
    print(f"\n[INFO] Merging results into Excel...")
    stats = merge_results(entries, reviewed)

    # 5. Clean up progress file
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)

    # 6. Print summary
    print(f"\n{'='*55}")
    print(f"  All done!")
    print(f"{'='*55}")
    total_excluded = stats.get('Excluded (non-target)', 0) + stats.get('Excluded (non-gene)', 0)
    print(f"  Confirmed:          {stats['Confirmed']}")
    print(f"  Weak association:   {stats['Weak association']}")
    print(f"  Excluded (non-target): {stats.get('Excluded (non-target)', 0)}")
    print(f"  Excluded (non-gene):   {stats.get('Excluded (non-gene)', 0)}")
    print(f"  Total excluded:     {total_excluded}")
    print(f"  Final retained:     {stats['Confirmed'] + stats['Weak association']}")
    print(f"{'='*55}")
    print(f"  Output file: {OUTPUT_XLSX}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
